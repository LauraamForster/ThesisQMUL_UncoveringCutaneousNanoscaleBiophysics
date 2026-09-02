#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Sep 26 08:57:53 2025

@author: lauraforster
"""

import numpy as np
import matplotlib.pyplot as plt
import pandas as pd  
from matplotlib.backends.backend_pdf import PdfPages
import ramanspy
from pathlib import Path
from sklearn.decomposition import PCA as _SKPCA
from sklearn.decomposition import TruncatedSVD as _SVD
from collections import defaultdict
from scipy.signal import savgol_filter, find_peaks
from itertools import cycle
from math import ceil
import warnings
from statsmodels.tools.sm_exceptions import ConvergenceWarning
# ---------------------------------------------------------------------------------------------------------
# -----------------------------------------------Initialisation--------------------------------------------
# ---------------------------------------------------------------------------------------------------------
"""
Raman linescan initialisation utilities.

This module prepares a working dictionary of samples and spectra, and (optionally) plots
spectra for a chosen sample. Typical flow:
1) Types(Type) -> returns Subtypes and default plot styles for that Type.
2) read_Samplemanifest(...) / read_Peakmanifest(...) -> load manifests.
3) CreateDict(sample_manifest_df, Subtypes) -> build data_dict index of samples to process.
4) readindata(DataDir, data_dict) -> read FP/EXT TXT files (X, Y, Wave, Intensity) into data_dict.
5) SplitSpectra(data_dict, Colours, Linestyles, SN, step) -> split each TXT into per-(X,Y)
   spectra, drop raw DataFrames to save memory, and optionally plot all spectra for sample SN.

Notes:
- SplitSpectra groups by the X,Y found in the TXT files; it does NOT assign skin layers from
  the Excel manifest. If layer tagging is needed, add a mapping step from (X,Y) to layer.
- EXTENDED_SUFFIX can be used to build extended file names consistently.

Functions:
- Types(Type): Return (Subtypes, Colours, Linestyles) for the given high-level Type.
- read_Samplemanifest(path): Load Excel sample manifest -> DataFrame.
- read_Peakmanifest(path): Load CSV peak/feature manifest -> DataFrame.
- CreateDict(sample_manifest_df, Subtypes): From the manifest, create a data_dict keyed by
  sample number with Type/Subtype metadata (filtered to allowed Subtypes).
- readindata(DataDir, data_dict): For each sample, read '{sample}A_linescan1.txt' and
  '{sample}A_linescan1_extended.txt' into data_dict['FP'] / data_dict['EXT'] with ffilled X,Y.
- SplitSpectra(data_dict, Colours, Linestyles, SN, step): Convert FP/EXT DataFrames into
  dicts of spectra per (X,Y) coordinate; optionally plot all spectra for sample SN, thinning
  by 'step' (plot every nth spectrum).
"""

def Types(Type):
    if Type == 'AP1':
        Subtypes = ["TS", "VH", "CL", "AC"]
        Colours = {"TS": "blue", "VH": "green",  "AC": "red", "CL": "peru"}
        Linestyles = {"TS": "-", "VH": "--",  "CL": "--", "AC": ":"}
    elif Type == 'WOUND':
        Subtypes = ["CT", "PBS", "D7", "D10" ,"D14", "D21"]
        Colours = {"CT": "grey","PBS": "blue", "D7": "tomato",  "D10": "mediumorchid", "D14": "royalblue", "D21": "mediumseagreen"}
        Linestyles = {"CT": "-", "PBS": "-", "D7": "--",  "D10": "--", "D14": ":", "D21": "-"}
        # Subtypes = ["CT", "D7", "D14", ]
        # Colours = {"CT": "blue", "D7": "green",   "D14": "peru"}
        # Linestyles = {"CT": "-", "D7": "--", "D14": ":"}

    elif Type == 'BLEO':
        Subtypes = ["PBS", "2W", "4W_5R"]
        Colours = {"PBS": "blue", "2W": "green", "4W_5R": "peru"}
        Linestyles = {"PBS": "-", "2W": "--", "4W_5R": ":"}
    elif Type == 'BLEO_MET':
        Subtypes = ["PBS_MET", "BM_MET", "4W_MET"]
        Colours = {"PBS_MET": "blue", "BM_MET": "green", "4W_MET": "peru"}
        Linestyles = {"PBS_MET": "-", "BM_MET": "--", "4W_MET": ":"}
    elif Type == 'BLEO_OKN':
        Subtypes = ["PBS_OKN",  "4W_OKN"]
        Colours = {"PBS_OKN": "blue", "4W_OKN": "peru"}
        Linestyles = {"PBS_OKN": "-", "4W_OKN": ":"}
    
    return Subtypes, Colours, Linestyles

def count_linescan_points(data_dict):
    """
    Make a table of how many spectra (positions) each repeat has:
      - N_total: number of spectra along the processed line (prefers *Treated*, then raw)
      - N_dermis: number of spectra in the Dermis segment (after TrimRegion)
    Counts are reported per sample and per region (FP/EXT where available).

    Returns
    -------
    df : pd.DataFrame with columns:
         ["Sample", "Subtype", "Region", "N_total", "N_dermis"]
    Also stores a copy in data_dict["_counts"]["per_region"].
    """
    rows = []
    for sample_num, sdict in data_dict.items():
        if not isinstance(sdict, dict):
            continue
        subtype = sdict.get("Subtype", None)

        for reg in ("FP", "EXT"):
            # Prefer processed (analysis window) if present; otherwise fall back.
            total_dict = (
                sdict.get(f"{reg}_Spectra_Treated") or
                sdict.get(f"{reg}_Spectra_Treated_Full") or
                sdict.get(f"{reg}_Spectra")
            )
            dermis_dict = sdict.get(f"{reg}_Spectra_Treated_Dermis")

            if (total_dict is None) and (dermis_dict is None):
                # nothing for this region in this sample—skip
                continue

            n_total  = len(total_dict)  if isinstance(total_dict, dict)  else 0
            n_dermis = len(dermis_dict) if isinstance(dermis_dict, dict) else 0

            rows.append({
                "Sample": str(sample_num),
                "Subtype": subtype,
                "Region": reg,
                "N_total": int(n_total),
                "N_dermis": int(n_dermis),
            })

    df = pd.DataFrame(rows).sort_values(["Subtype", "Sample", "Region"], ignore_index=True)

    # store for later use
    data_dict.setdefault("_counts", {})["per_region"] = df
    return df, data_dict


def read_Samplemanifest(manifest_path):
    sample_df = pd.read_excel(manifest_path)
    return sample_df

def read_Peakmanifest(manifest_path, sheet_name="Paper"):
    manifest_path = Path(manifest_path)

    if manifest_path.suffix.lower() in [".xlsx", ".xls"]:
        peak_df = pd.read_excel(manifest_path, sheet_name=sheet_name)
    elif manifest_path.suffix.lower() == ".csv":
        peak_df = pd.read_csv(manifest_path)
    else:
        raise ValueError(f"Unsupported peak manifest format: {manifest_path.suffix}")

    peak_df.columns = [str(c).strip() for c in peak_df.columns]
    return peak_df

def BuildAssignmentsFromManifest(peak_manifest_df, component_colours=None):
    """
    Build assignment dictionary from peak manifest.

    Expected columns:
      - Region
      - Component
      - Position

    Returns
    -------
    assignments : dict
        {Component: [positions]}
    assign_colours : dict
        {Component: colour}
    """
    required = {"Component", "Position"}
    missing = required.difference(peak_manifest_df.columns)
    if missing:
        raise ValueError(f"Peak manifest missing required columns: {sorted(missing)}")

    df = peak_manifest_df.copy()
    df["Component"] = df["Component"].astype(str).str.strip()
    df["Position"] = pd.to_numeric(df["Position"], errors="coerce")
    df = df.dropna(subset=["Component", "Position"])

    assignments = {}
    for comp, g in df.groupby("Component", sort=False):
        assignments[comp] = sorted(g["Position"].astype(float).unique().tolist())

    # default colours if not provided
    default_cycle = plt.rcParams["axes.prop_cycle"].by_key().get("color", [])
    if not default_cycle:
        default_cycle = ["tab:blue", "tab:orange", "tab:green", "tab:red", "tab:purple", "tab:brown", "tab:pink"]

    assign_colours = {}
    for i, comp in enumerate(assignments.keys()):
        if component_colours and comp in component_colours:
            assign_colours[comp] = component_colours[comp]
        else:
            assign_colours[comp] = default_cycle[i % len(default_cycle)]

    return assignments, assign_colours

def CreateDict(sample_manifest_df, Subtypes):
    """
    Build base data_dict from manifest.
    Stores direction per sample for later binning.
    """

    data_dict = {}

    for _, row in sample_manifest_df.iterrows():
        sample_num = str(int(row["Sample Number"])).strip()
        subtype = str(row["TYPE"]).strip()

        if subtype in Subtypes:
            data_dict[sample_num] = {
                "Type": subtype,
                "Subtype": subtype,
                "direction": str(row["direction"]).strip().lower()  # <-- NEW
            }

    return data_dict
    
def readindata(DataDir, data_dict):
    
    # 1. From the dictionary we know which sample numbers we have and their group (subtype)
    for sample_num in data_dict.keys():
        subtype = data_dict[sample_num]["Subtype"]
        base_path = DataDir / subtype / f"{sample_num}A"
    
    # 2. Find each .txt file 
        # Format is DataDir path +/subtype +/{samplenumber}A_linescan1.txt and +/{samplenumber}A_linescan1_extended.txt
        fp_path = base_path.with_name(f"{sample_num}A_linescan1.txt")
        ext_path = base_path.with_name(f"{sample_num}A_linescan1_extended.txt")
        
    # 3. Read in both the fingerprint and extended .txt files as 4 columns and save into the dictionary under the sample number
        for path, key in [(fp_path, "FP"), (ext_path, "EXT")]:
            try:
                # Robust read: whitespace-sep, skip comment lines
                df = pd.read_csv(path, sep=r"\s+", comment="#", header=None, names=["X", "Y", "Wave", "Intensity"])
                
                # Forward-fill X and Y to replicate Raman structure
                df[["X", "Y"]] = df[["X", "Y"]].ffill()

                # Drop header rows or junk rows
                df = df[df["Wave"].apply(lambda x: isinstance(x, (float, int)))]

                data_dict[sample_num][key] = df

            except FileNotFoundError:
                print(f"{key} file not found for sample {sample_num}: {path}")

    return data_dict

def SplitSpectra(data_dict, Colours, Linestyles, SN, step):

    for sample_num in data_dict.keys():
        for region in ["FP", "EXT"]:
            key_name = f"{region}_Spectra"
            spectra = {}

            if region not in data_dict[sample_num]:
                continue

            df = data_dict[sample_num][region]

            # Group rows by (X, Y) coordinate — each group is one spectrum
            grouped = df.groupby(["X", "Y"], sort=False)
            for (x, y), group in grouped:
                wave = group["Wave"].values
                intensity = group["Intensity"].values
                spectra[(x, y)] = {"Wave": wave, "Intensity": intensity}

            data_dict[sample_num][key_name] = spectra
# Deleting Steps
        # Clean up raw DataFrames to save space
        # we can then delete the fp_df and ext_df from the dictionary and just store all the spectra under their own sample number key
        data_dict[sample_num].pop("FP", None)
        data_dict[sample_num].pop("EXT", None)

# Plotting steps

        if SN is not None and str(SN) == sample_num:
            subtype = data_dict[sample_num].get("Subtype")
            colour = Colours.get(subtype, None) if Colours else None
            linestyle = Linestyles.get(subtype, None) if Linestyles else None
            for region in ["FP_Spectra", "EXT_Spectra"]:
                if region not in data_dict[sample_num]:
                    continue

                plt.figure(figsize=(10, 6))
                spectra_dict = data_dict[sample_num][region]

                for i, ((x, y), spectrum) in enumerate(spectra_dict.items()):
                    if step is None or i % step == 0:
                        x,y = np.round(x,2), np.round(y,2)
                        label = f"x={x}, y={y}"
                        plt.plot(spectrum["Wave"], spectrum["Intensity"], label=label, color=colour, linestyle=linestyle)

                plt.title(f"{region} spectra for Sample {sample_num}")
                plt.xlabel("Wavenumber (cm⁻¹)")
                plt.ylabel("Intensity")
                plt.legend(fontsize="x-small", loc="upper right", ncol=2)
                plt.tight_layout()
                plt.show()

    return data_dict

# ---------------------------------------------------------------------------------------------------------
# -----------------------------------------------Preprocessing---------------------------------------------
# ---------------------------------------------------------------------------------------------------------
_CANONICAL_ORDER = ["despike", "smooth", "baseline", "normalise"]

def _minmax_for_plot(y):
    y = np.asarray(y, dtype=float)
    # robust min/max using percentiles to avoid spikes dominating
    lo, hi = np.percentile(y, 1), np.percentile(y, 99)
    if not np.isfinite(lo) or not np.isfinite(hi) or hi <= lo:
        lo, hi = np.min(y), np.max(y)
    if hi <= lo:
        return np.zeros_like(y)
    return (y - lo) / (hi - lo)

def _normalize_steps(preprocess_list):
    """Return a set of valid steps, lowercased, filtered to known names."""
    valid = {s.lower() for s in preprocess_list or []}
    unknown = valid.difference(_CANONICAL_ORDER)
    if unknown:
        print(f"Warning: ignoring unknown steps: {sorted(unknown)}")
    return [s for s in _CANONICAL_ORDER if s in valid]  # keep canonical order

def _build_region_steps(full_rng, crop_rng, selected_steps):
    """Build callables and titles for the region, with crop at start and end."""
    cropper_full  = ramanspy.preprocessing.misc.Cropper(region=full_rng)
    despike       = ramanspy.preprocessing.despike.WhitakerHayes()
    savgol        = ramanspy.preprocessing.denoise.SavGol(window_length=11, polyorder=3)
    # baseline        = ramanspy.preprocessing.baseline.ASLS()
    baseline = ramanspy.preprocessing.baseline.ASLS(p=0.01,lam=1e4)
    auc           = ramanspy.preprocessing.normalise.AUC(pixelwise=True)
    cropper_final = ramanspy.preprocessing.misc.Cropper(region=crop_rng)

    step_map = {
        "despike":    (lambda sc: despike.apply(sc),    "After despike (WhitakerHayes)"),
        "smooth":     (lambda sc: savgol.apply(sc),     "After smoothing (SavGol)"),
        "baseline":   (lambda sc: baseline.apply(sc),     "After baseline (ASLS)"),
        "normalise":  (lambda sc: auc.apply(sc),        "After normalisation (AUC)"),
    }

    # full pipeline (no final crop) used for stored *_Treated_Full and overlay middle trace
    step_fns_full  = [lambda sc: cropper_full.apply(sc)]
    titles_full    = ["After crop (full range)"]
    for name in selected_steps:
        fn, title = step_map[name]
        step_fns_full.append(fn)
        titles_full.append(title)

    # final-cropped pipeline (adds final crop)
    step_fns_final = step_fns_full + [lambda sc: cropper_final.apply(sc)]
    titles_final   = titles_full + ["After final crop (analysis window)"]

    return step_fns_full, titles_full, step_fns_final, titles_final

def _style_for(label):
    """Consistent, visible styles by label substring."""
    lab = label.lower()
    if "raw" in lab:
        return dict(ls="-", lw=1.3, alpha=0.95, zorder=1)
    if "crop (full" in lab:
        return dict(ls="None", marker="o", ms=2.2, alpha=0.9, zorder=5, color="tab:orange")
    if "despike" in lab:
        return dict(ls="--", lw=1.5, alpha=0.95, zorder=3, color="tab:green")
    if "smoothing" in lab:
        return dict(ls="-", lw=2.0, alpha=0.95, zorder=6, color="tab:red")
    if "baseline" in lab:
        return dict(ls=":", lw=1.6, alpha=0.9, zorder=4, color="tab:purple")
    if "normalisation" in lab:
        return dict(ls="-.", lw=1.4, alpha=0.9, zorder=2, color="tab:gray")
    if "final crop" in lab:
        return dict(ls="None", marker=".", ms=2.6, alpha=0.95, zorder=7, color="saddlebrown")
    # fallback
    return dict(ls="-", lw=1.2, alpha=0.9, zorder=2)

def _pipeline_debug_fig(sc_raw, step_fns, titles, suptitle="", norm_band=None):
    """
    Multi-panel debug figure (9 panels total):

    - Still RUNS all step_fns (processing is faithful)
    - Hides old:
        * "After normalisation (AUC)"
        * "After final crop (analysis window)"
    - Inserts new:
        5) After AUC normalisation (FP_crop / analysis window)
        6) After normalisation to band (norm_band)
        7) Overlay of both

    Also:
      - baseline preview panel: estimated baseline + previous
      - "After baseline" panel: y=0 line + negative area fraction annotation
    """
    

    def _auc(x, y):
        x = np.asarray(x, float)
        y = np.asarray(y, float)
        if x.size < 2 or y.size < 2:
            return np.nan
        return float(np.trapezoid(y, x))

    def _crop_xy(x, y, region):
        if region is None:
            return x, y
        lo, hi = float(region[0]), float(region[1])
        x = np.asarray(x, float)
        y = np.asarray(y, float)
        m = (x >= lo) & (x <= hi) & np.isfinite(x) & np.isfinite(y)
        return x[m], y[m]

    def _is_old_auc_panel(title):
        t = str(title).lower()
        return ("normalisation" in t and "auc" in t) or ("normalization" in t and "auc" in t)

    def _is_old_final_crop_panel(title):
        t = str(title).lower()
        return ("final crop" in t) or ("analysis window" in t)

    def _negative_area_fraction(y):
        y = np.asarray(y, float)
        if y.size == 0 or not np.all(np.isfinite(y)):
            return np.nan
        denom = np.sum(np.abs(y)) + 1e-12
        return float(np.sum(np.abs(y[y < 0])) / denom)

    # stages: (title, x, y, y_overlay_or_None)
    stages = []

    x_prev = np.asarray(sc_raw.spectral_axis, float)
    y_prev = np.asarray(sc_raw.spectral_data, float)
    stages.append(("0 – Raw", x_prev, y_prev, None))

    sc_prev = sc_raw

    # capture baseline-corrected + final-crop output (even if we hide their panels)
    x_basecorr = y_basecorr = None
    x_final = y_final = None

    for i, (fn, name) in enumerate(zip(step_fns, titles), start=1):
        sc_curr = fn(sc_prev)
        x = np.asarray(sc_curr.spectral_axis, float)
        y = np.asarray(sc_curr.spectral_data, float)

        t_clean = str(name).strip()
        t_lc = t_clean.lower()
        is_baseline = ("baseline" in t_lc)

        if _is_old_final_crop_panel(t_clean):
            x_final, y_final = x, y

        hide_this_panel = _is_old_auc_panel(t_clean) or _is_old_final_crop_panel(t_clean)

        if is_baseline:
            # align previous onto current x (avoid broadcast issues)
            if (len(x_prev) != len(x)) or (not np.allclose(x_prev, x, rtol=0, atol=1e-9)):
                y_prev_aligned = np.interp(x, x_prev, y_prev)
            else:
                y_prev_aligned = y_prev

            baseline_est = y_prev_aligned - y

            stages.append((f"{i} – Baseline (preview)", x, baseline_est, y_prev_aligned))
            stages.append((f"{i} – {t_clean}", x, y, None))

            x_basecorr, y_basecorr = x, y

        else:
            if not hide_this_panel:
                stages.append((f"{i} – {t_clean}", x, y, None))

        sc_prev, x_prev, y_prev = sc_curr, x, y

    # insert new panels right after "After baseline" if present
    insert_at = None
    for k, (t, *_rest) in enumerate(stages):
        if "after baseline" in t.lower():
            insert_at = k + 1
            break

    if (x_basecorr is not None) and (y_basecorr is not None) and (x_final is not None) and (y_final is not None):
        # align baseline-corrected onto final axis
        if (len(x_basecorr) != len(x_final)) or (not np.allclose(x_basecorr, x_final, rtol=0, atol=1e-9)):
            y_base_on_final = np.interp(x_final, x_basecorr, y_basecorr)
        else:
            y_base_on_final = y_basecorr

        # option 1: AUC normalise over FP_crop (final window)
        auc_crop = _auc(x_final, y_base_on_final)
        denom1 = auc_crop if (np.isfinite(auc_crop) and abs(auc_crop) > 0) else 1.0
        y_opt1 = y_base_on_final / denom1

        # option 2: AUC normalise over norm_band within final window
        if norm_band is not None:
            xb, yb = _crop_xy(x_final, y_base_on_final, norm_band)
            auc_band = _auc(xb, yb)
        else:
            auc_band = np.nan
        denom2 = auc_band if (np.isfinite(auc_band) and abs(auc_band) > 0) else 1.0
        y_opt2 = y_base_on_final / denom2

        new_panels = [
            ("5 – After AUC normalisation (FP_crop)", x_final, y_opt1, None),
            (f"6 – After normalisation to band {norm_band}", x_final, y_opt2, None),
            ("7 – Overlay: FP_crop AUC vs band-normalised", x_final, y_opt1, y_opt2),
        ]

        if insert_at is None:
            stages.extend(new_panels)
        else:
            stages[insert_at:insert_at] = new_panels

    # layout (expect 9 panels, but robust anyway)
    n = len(stages)
    ncols = 3 if n >= 3 else n
    nrows = int(ceil(n / ncols))
    fig, axes = plt.subplots(nrows, ncols, figsize=(5.3 * ncols, 3.6 * nrows), squeeze=False)

    for idx, ax in enumerate(axes.flat):
        if idx >= n:
            ax.axis("off")
            continue

        title_i, x, y, y2 = stages[idx]
        ax.plot(x, y, lw=1.6)
        if y2 is not None:
            ax.plot(x, y2, lw=1.6)

        # baseline-corrected diagnostic panel
        if "after baseline" in title_i.lower():
            ax.axhline(0, color="k", lw=1.0, alpha=0.7)
            naf = _negative_area_fraction(y)
            ax.text(
                0.02, 0.95, f"Neg. area frac ≈ {naf:.3f}",
                transform=ax.transAxes, va="top", ha="left", fontsize=9,
                bbox=dict(boxstyle="round,pad=0.25", fc="w", ec="0.7", alpha=0.85)
            )

        ax.set_title(title_i, fontsize=11)
        ax.set_xlabel("Wavenumber (cm⁻¹)")
        ax.set_ylabel("Intensity (a.u.)")
        ax.grid(True, linestyle=":", alpha=0.25)

        if "baseline (preview)" in title_i.lower():
            ax.legend(["baseline", "previous"], fontsize=9, frameon=True, framealpha=0.8)

        if "overlay: fp_crop auc vs band-normalised" in title_i.lower():
            ax.legend(["FP_crop AUC", "Band-normalised"], fontsize=9, frameon=True, framealpha=0.8)

    if suptitle:
        fig.suptitle(suptitle, y=0.995, fontsize=12)

    fig.tight_layout()
    return fig

def _overlay_fig(raw_wave, raw_int, pf_wave, pf_int, pc_wave, pc_int, title_suffix):
    """Overlay raw vs processed(no crop) vs processed+cropped with visible styles."""
    plt.figure(figsize=(10, 6))
    plt.plot(raw_wave, _minmax_for_plot(raw_int), label="Raw (scaled)", **_style_for("raw"))
    plt.plot(pf_wave,  _minmax_for_plot(pf_int),  label="Processed (no crop, scaled)", ls="--", lw=1.6, alpha=0.95, zorder=3)
    plt.plot(pc_wave,  _minmax_for_plot(pc_int),  label="Processed + Cropped (scaled)", ls="None", marker=".", ms=2.6, alpha=0.95, zorder=4)
    plt.title(f"Pipeline overlay — {title_suffix}")
    plt.xlabel("Wavenumber (cm⁻¹)")
    plt.ylabel("Scaled intensity (0–1)")
    plt.legend()
    plt.tight_layout()
    return plt.gcf()


def _normalise_sc(sc_in, mode, band):
    """
    mode:
      - "crop": divide by AUC over the current window
      - "band": divide by AUC over `band` (within current window); falls back to crop if band invalid
    """
    import numpy as np
    import ramanspy

    w = np.asarray(sc_in.spectral_axis, float)
    y = np.asarray(sc_in.spectral_data, float)
    if w.size < 2 or y.size < 2 or not np.all(np.isfinite(y)):
        return sc_in

    denom = np.nan
    if mode == "band" and band is not None:
        lo, hi = float(band[0]), float(band[1])
        m = (w >= lo) & (w <= hi) & np.isfinite(w) & np.isfinite(y)
        if np.any(m):
            denom = float(np.trapezoid(y[m], w[m]))

    if not (np.isfinite(denom) and denom != 0):
        denom = float(np.trapezoid(y, w))

    if not (np.isfinite(denom) and denom != 0):
        denom = 1.0

    sc_out = ramanspy.SpectralContainer(y / denom, w)
    if hasattr(sc_in, "metadata"):
        try:
            sc_out.metadata = sc_in.metadata
        except Exception:
            pass
    return sc_out

def TreatSpectra(data_dict, Save_folder, Preprocess,
                 FP_full, FP_crop, EXT_full, EXT_crop,
                 colours, linestyles, plotall_treat, treatmentorder, normalisation,
                 SN, step,
                 FP_band=None, EXT_band=None):
    """
    Updated wounding-style TreatSpectra that matches your current Bovine logic.

    plotall_treat: "screen" | "pdf" | "None"/"false"
      - "screen": ONLY pipeline debug figure
      - "pdf": pipeline debug + overlay figure to ONE PDF PER SAMPLE
      - "None"/"false": no plots

    normalisation: "crop" | "band"
      - controls what is stored into *_Treated (downstream-used)
      - *_Treated_Full remains the processed full-range output (no final crop)
    """
    from pathlib import Path
    import numpy as np
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages
    import ramanspy

    # ---- parse plot mode ----
    mode = str(plotall_treat).strip().lower()
    if mode in ("none", "false", "0", ""):
        plot_mode = "false"
    elif mode in ("screen", "pdf"):
        plot_mode = mode
    else:
        plot_mode = "screen" if bool(plotall_treat) else "false"

    order_mode = str(treatmentorder).strip().lower()
    if order_mode not in ("before", "after"):
        print(f"[TreatSpectra] Unknown treatmentorder='{treatmentorder}', using 'before'.")
        order_mode = "before"

    norm_mode = str(normalisation).strip().lower()
    if norm_mode not in ("crop", "band"):
        print(f"[TreatSpectra] Unknown normalisation='{normalisation}', using 'crop'.")
        norm_mode = "crop"

    selected = _normalize_steps(Preprocess)

    # ---- plotting filter only ----
    if SN is None:
        plot_samples = None
    else:
        plot_samples = set(SN) if isinstance(SN, (list, tuple, set)) else {SN}

    # ---- output dir for pdf ----
    outdir = None
    if plot_mode == "pdf":
        outdir = Path(Save_folder) / "Outputs"
        outdir.mkdir(parents=True, exist_ok=True)

    for sample_num, sample_data in data_dict.items():
        treated_full = {"FP_Spectra": {}, "EXT_Spectra": {}}
        treated_crop = {"FP_Spectra": {}, "EXT_Spectra": {}}

        pdf_ctx = None
        if plot_mode == "pdf":
            pdf_path = outdir / f"Sample_{sample_num}_Treatedspectra.pdf"
            pdf_ctx = PdfPages(pdf_path)

        try:
            for region_key, region_label, full_rng, crop_rng, band_rng in [
                ("FP_Spectra",  "FP",  FP_full,  FP_crop,  FP_band),
                ("EXT_Spectra", "EXT", EXT_full, EXT_crop, EXT_band),
            ]:
                if region_key not in sample_data:
                    continue

                step_fns_full, titles_full, step_fns_final, titles_final = _build_region_steps(
                    full_rng, crop_rng, selected
                )
                crop_fn = step_fns_final[-1]  # analysis-window crop callable

                # stable order along the line
                items = sorted(
                    sample_data[region_key].items(),
                    key=lambda kv: (round(kv[0][0], 9), round(kv[0][1], 9))
                )

                for idx, ((x, y), spec) in enumerate(items):
                    plot_allowed_sample = (plot_samples is None) or (sample_num in plot_samples)
                    plot_step_ok = (step in (None, 0, 1)) or (isinstance(step, int) and step > 1 and idx % step == 0)
                    plot_this = (plot_mode != "false") and plot_allowed_sample and plot_step_ok

                    wave = np.asarray(spec["Wave"], float)
                    inten = np.asarray(spec["Intensity"], float)

                    # ensure increasing x
                    if wave.size and np.any(np.diff(wave) < 0):
                        o = np.argsort(wave)
                        wave = wave[o]
                        inten = inten[o]

                    sc_raw = ramanspy.SpectralContainer(inten, wave)
                    sc_raw.metadata = {"x": x, "y": y, "index": idx}

                    if order_mode == "before":
                        sc_full = sc_raw
                        for fn in step_fns_full:
                            sc_full = fn(sc_full)

                        sc_final = crop_fn(sc_full)
                        sc_final_norm = _normalise_sc(sc_final, norm_mode, band_rng)
                        # sc_final_norm = sc_final

                        treated_full[region_key][(x, y)] = sc_full
                        treated_crop[region_key][(x, y)] = sc_final_norm

                        if plot_this:
                            xy = (f"Sample {sample_num} — {region_label} — i={idx} — "
                                  f"x={np.round(x,2)}, y={np.round(y,2)}")
                            fig_dbg = _pipeline_debug_fig(sc_raw, step_fns_final, titles_final, xy + " (before)", norm_band=band_rng)
                            if plot_mode == "pdf":
                                pdf_ctx.savefig(fig_dbg); plt.close(fig_dbg)
                            else:
                                plt.show(); plt.close(fig_dbg)

                            # overlay ONLY for pdf
                            if plot_mode == "pdf":
                                fig_ov = _overlay_fig(
                                    np.asarray(sc_raw.spectral_axis,float),  np.asarray(sc_raw.spectral_data,float),
                                    np.asarray(sc_full.spectral_axis,float), np.asarray(sc_full.spectral_data,float),
                                    np.asarray(sc_final_norm.spectral_axis,float), np.asarray(sc_final_norm.spectral_data,float),
                                    xy + " (before)"
                                )
                                pdf_ctx.savefig(fig_ov); plt.close(fig_ov)

                    else:  # after
                        sc_crop_first = crop_fn(sc_raw)
                        sc_proc = sc_crop_first
                        for fn in step_fns_full:
                            sc_proc = fn(sc_proc)

                        sc_final_norm = _normalise_sc(sc_proc, norm_mode, band_rng)
                        # sc_final_norm = sc_proc
                        
                        sc_full = sc_raw
                        for fn in step_fns_full:
                            sc_full = fn(sc_full)

                        treated_full[region_key][(x, y)] = sc_full
                        treated_crop[region_key][(x, y)] = sc_final_norm

                        if plot_this:
                            xy = (f"Sample {sample_num} — {region_label} — i={idx} — "
                                  f"x={np.round(x,2)}, y={np.round(y,2)}")
                            fig_dbg = _pipeline_debug_fig(
                                sc_raw, [crop_fn] + step_fns_full, ["Crop(ROI)"] + titles_full, xy + " (after)", norm_band=band_rng
                            )
                            if plot_mode == "pdf":
                                pdf_ctx.savefig(fig_dbg); plt.close(fig_dbg)
                            else:
                                plt.show(); plt.close(fig_dbg)

                            if plot_mode == "pdf":
                                fig_ov = _overlay_fig(
                                    np.asarray(sc_raw.spectral_axis,float),  np.asarray(sc_raw.spectral_data,float),
                                    np.asarray(sc_full.spectral_axis,float), np.asarray(sc_full.spectral_data,float),
                                    np.asarray(sc_final_norm.spectral_axis,float), np.asarray(sc_final_norm.spectral_data,float),
                                    xy + " (after)"
                                )
                                pdf_ctx.savefig(fig_ov); plt.close(fig_ov)

            if plot_mode == "pdf" and pdf_ctx is not None:
                cover = plt.figure(figsize=(10, 6))
                cover.text(0.5, 0.6, f"Sample {sample_num}", ha="center", va="center", fontsize=18)
                cover.text(0.5, 0.5, f"Order: {order_mode} | Normalisation: {norm_mode}",
                           ha="center", va="center", fontsize=12)
                cover.tight_layout()
                pdf_ctx.savefig(cover); plt.close(cover)

        finally:
            if plot_mode == "pdf" and pdf_ctx is not None:
                pdf_ctx.close()

        # save back (same key pattern as before)
        for rk in ["FP_Spectra", "EXT_Spectra"]:
            data_dict[sample_num][f"{rk}_Treated_Full"] = treated_full[rk]  # processed, no final crop
            data_dict[sample_num][f"{rk}_Treated"]      = treated_crop[rk]  # processed, cropped, AUC-normalised (crop/band)

    return data_dict

def TrimRegion(data_dict, sample_manifest_df, Colours, Linestyles, SN, step):
    for sample_num, sample_data in data_dict.items():
        manifest_row = sample_manifest_df[sample_manifest_df["Sample Number"].astype(str) == sample_num]
        if manifest_row.empty:
            print(f"Sample {sample_num} not found in manifest")
            continue

        dermis_x_start = manifest_row["dermis x"].values[0]
        dermis_y_start = manifest_row["dermis y"].values[0]
        dermis_x_end   = manifest_row["epi x"].values[0]
        dermis_y_end   = manifest_row["epi y"].values[0]
        dermis_length_reported = manifest_row["leng dermis"].values[0]

        direction = manifest_row["direction"].values[0].strip().lower()

        for region_key in ["FP_Spectra_Treated", "EXT_Spectra_Treated"]:
            if region_key not in sample_data:
                continue

            spectra_dict = sample_data[region_key]
            coords = np.array(list(spectra_dict.keys()))  # shape (N,2)

            # Trim dermis
            dist_start = np.linalg.norm(coords - np.array([dermis_x_start, dermis_y_start]), axis=1)
            dist_end   = np.linalg.norm(coords - np.array([dermis_x_end, dermis_y_end]), axis=1)

            idx_start = np.argmin(dist_start)
            idx_end   = np.argmin(dist_end)

            i1, i2 = sorted([idx_start, idx_end])
            trimmed_keys = [tuple(coords[i]) for i in range(i1, i2+1)]

            if direction == "back":
                trimmed_keys = list(reversed(trimmed_keys))

            trimmed_spectra = {k: spectra_dict[k] for k in trimmed_keys}
            data_dict[sample_num][f"{region_key}_Dermis"] = trimmed_spectra

            # length check
            if abs(len(trimmed_spectra) - dermis_length_reported) > 2:
                print(f"Error! Mismatch for {sample_num} {region_key}: "
                      f"manifest={int(dermis_length_reported)} vs trimmed={len(trimmed_spectra)}")

        if SN is not None and str(SN) == sample_num:
            subtype = data_dict[sample_num].get("Subtype")
            colour = Colours.get(subtype, None) if Colours else None
            linestyle = Linestyles.get(subtype, None) if Linestyles else None
            for region_key in ["FP_Spectra_Treated_Dermis", "EXT_Spectra_Treated_Dermis"]:
                if region_key not in data_dict[sample_num]:
                    continue
                plt.figure(figsize=(10, 6))
                spectra_dict = data_dict[sample_num][region_key]
                for i, ((x, y), spectrum) in enumerate(spectra_dict.items()):
                    if step is None or i % step == 0:
                        plt.plot(spectrum.spectral_axis, spectrum.spectral_data,
                                 label=f"x={np.round(x,2)}, y={np.round(y,2)}",
                                 color=colour, linestyle=linestyle)
                plt.title(f"{region_key} spectra for Sample {sample_num}")
                plt.xlabel("Wavenumber (cm⁻¹)")
                plt.ylabel("Intensity")
                plt.legend(fontsize="x-small", loc="upper right", ncol=2)
                plt.tight_layout()
                plt.show()

    return data_dict

def BinData(data_dict, NBins, Colours, Linestyles, ST=None, region=None, Binno=None, SN=None):
    """
    Bin trimmed dermis spectra for each subtype and region into NBins along scan order.
    - Preserves the insertion order from TrimRegion (no sorting).
    - Interpolates spectra in each bin to a common wave axis before averaging.
    - Stores per-sample bins back into data_dict under keys "FP_Bins" and "EXT_Bins".
    Returns:
        binned_avg[subtype][region_short] -> list of dicts {"Wave", "Intensity"} per bin
    """
    subtype_samples = {}
    for sample_num, sample_data in data_dict.items():
        subtype = sample_data.get("Subtype")
        if subtype is None:
            continue
        subtype_samples.setdefault(subtype, []).append(sample_num)

    binned_avg = {}
    for subtype, samples in subtype_samples.items():
        print(f"Processing subtype: {subtype}")
        binned_avg[subtype] = {"FP": [], "EXT": []}

        for region_key in ["FP_Spectra_Treated_Dermis", "EXT_Spectra_Treated_Dermis"]:
            region_short = region_key.split("_")[0]  # "FP" or "EXT"
            samples_binned = []   # list of [bin0, bin1, ...] per-sample

            for sample_num in samples:
                if region_key not in data_dict[sample_num]:
                    continue

                spectra_dict = data_dict[sample_num][region_key]
                keys_order = list(spectra_dict.keys())  # preserve scan order
                n_points = len(keys_order)
                if n_points == 0:
                    continue

                N_eff = min(NBins, n_points)
                index_bins = np.array_split(np.arange(n_points), N_eff)

                ref_axis = np.asarray(spectra_dict[keys_order[0]].spectral_axis, dtype=float)

                bin_averages = []
                for idxs in index_bins:
                    if len(idxs) == 0:
                        bin_averages.append(None)
                        continue

                    bin_intensities = []
                    for idx in idxs:
                        sc = spectra_dict[keys_order[idx]]
                        wav = np.asarray(sc.spectral_axis, dtype=float)
                        inten = np.asarray(sc.spectral_data, dtype=float)
                        if wav.shape != ref_axis.shape or not np.allclose(wav, ref_axis, rtol=0, atol=1e-9):
                            inten = np.interp(ref_axis, wav, inten)
                        bin_intensities.append(inten)

                    avg_intensity = np.mean(np.vstack(bin_intensities), axis=0)
                    bin_averages.append({"Wave": ref_axis, "Intensity": avg_intensity})

                samples_binned.append(bin_averages)

                # --- Save bins back to data_dict per sample ---
                data_dict[sample_num][f"{region_short}_Bins"] = bin_averages

            # Merge across samples (per bin) for subtype-level average
            if not samples_binned:
                binned_avg[subtype][region_short] = []
                continue

            N_bins_final = len(samples_binned[0])
            bins_merged = []
            for b in range(N_bins_final):
                bin_stack = [s[b]["Intensity"] for s in samples_binned if s[b] is not None]
                if not bin_stack:
                    continue
                bin_stack = np.vstack(bin_stack)
                mean_intensity = np.mean(bin_stack, axis=0)
                wave_axis = None
                for s in samples_binned:
                    if s[b] is not None:
                        wave_axis = s[b]["Wave"]
                        break
                bins_merged.append({"Wave": wave_axis, "Intensity": mean_intensity})

            binned_avg[subtype][region_short] = bins_merged

    # --- Plotting (optional) ---
    if ST is not None or SN is not None:
        plt.figure(figsize=(8, 5))

        # Subtype-level plotting
        if ST is not None:
            if isinstance(ST, str):
                ST = [ST]
            if isinstance(Binno, int):
                Binno = [Binno]

            for st in ST:
                if st not in binned_avg:
                    print(f"Warning: subtype {st} not found in binned data.")
                    continue
                if region not in binned_avg[st]:
                    print(f"Warning: region {region} not found for subtype {st}.")
                    continue
                bins = binned_avg[st][region]
                colour = Colours.get(st, None) if Colours else None
                linestyle = Linestyles.get(st, None) if Linestyles else None

                for b in Binno:
                    if b < 0 or b >= len(bins):
                        print(f"Warning: bin number {b} out of range for subtype {st} region {region}.")
                        continue
                    spectrum = bins[b]
                    plt.plot(spectrum["Wave"], spectrum["Intensity"],
                             label=f"{st} bin {b}", color=colour, linestyle=linestyle)

        # Per-sample plotting
        if SN is not None and str(SN) in data_dict:
            for reg_short in ["FP", "EXT"]:
                key = f"{reg_short}_Bins"
                if key not in data_dict[str(SN)]:
                    continue
                bins = data_dict[str(SN)][key]
                for i, spectrum in enumerate(bins):
                    if spectrum is None:
                        continue
                    plt.plot(spectrum["Wave"], spectrum["Intensity"],
                             label=f"Sample {SN} {reg_short} bin {i}")

        plt.title(f"Binned Spectrum (Region={region})")
        plt.xlabel("Wavenumber (cm⁻¹)")
        plt.ylabel("Intensity")
        plt.legend()
        plt.tight_layout()
        plt.show()

    return data_dict, binned_avg

def PlotRegionAverageSpectra(
    data_dict,
    TypestoPlot,
    region="dermis",
    use_FP=True,
    use_treated=True,
    AveragebyType=False,
    show_error=True,
    error_alpha=0.50,
    err_mode="sem",
    axvlines=None,
    xlim=None,
    ylim=None,
    verbose=False
):

    region = str(region).strip().lower()
    typelist = [str(t).strip() for t in (TypestoPlot if isinstance(TypestoPlot, (list, tuple, set)) else [TypestoPlot])]
    typelist_set = set(typelist)

    reg_short = "FP" if use_FP else "EXT"
    treated_tag = "Treated" if use_treated else "Treated_Full"
    region_key = f"{reg_short}_Spectra_{treated_tag}_{region.capitalize()}"

    def _get_subtype(sample_data):
        return str(sample_data.get("Subtype", "")).strip()

    def _extract_region_spectra(sample_data):
        sd = sample_data.get(region_key, None)
        return sd if isinstance(sd, dict) and len(sd) else None

    def _align_stack(spectra_dict):
        keys = list(spectra_dict.keys())
        sc0 = spectra_dict[keys[0]]
        x0 = np.asarray(sc0.spectral_axis, float)

        Ys = []
        for k in keys:
            sc = spectra_dict[k]
            x = np.asarray(sc.spectral_axis, float)
            y = np.asarray(sc.spectral_data, float)
            if (x.shape != x0.shape) or (not np.allclose(x, x0, rtol=0, atol=1e-9)):
                y = np.interp(x0, x, y)
            Ys.append(y)

        Y = np.vstack(Ys) if Ys else np.empty((0, x0.size))
        return x0, Y

    def _mean_and_err(Y):
        if Y.size == 0:
            return None, None, 0
        y_mean = np.nanmean(Y, axis=0)
        n_specs = Y.shape[0]
        if n_specs <= 1:
            y_err = np.zeros_like(y_mean)
        else:
            y_std = np.nanstd(Y, axis=0, ddof=1)
            y_err = y_std if err_mode.lower() == "std" else (y_std / np.sqrt(n_specs))
        return y_mean, y_err, n_specs

    def _plot_peak_lines(ax):
        if axvlines:
            for i, xv in enumerate(axvlines):
                ax.axvline(
                    xv, color="k", linestyle="--", alpha=0.35, linewidth=1,
                    label="Reference peaks" if i == 0 else None
                )

    curves = []

    if AveragebyType:
        pool = {}
        for sample_num, sample_data in data_dict.items():
            st = _get_subtype(sample_data)
            if st not in typelist_set:
                continue

            sd = _extract_region_spectra(sample_data)
            if sd is None:
                if verbose:
                    print(f"[PlotRegionAverageSpectra] Missing {region_key} for sample {sample_num} ({st})")
                continue

            x0, Y = _align_stack(sd)
            if st not in pool:
                pool[st] = {"x": x0, "Ys": [], "sample_nums": set()}
            else:
                x_ref = pool[st]["x"]
                if (x0.shape != x_ref.shape) or (not np.allclose(x0, x_ref, rtol=0, atol=1e-9)):
                    Y = np.vstack([np.interp(x_ref, x0, row) for row in Y]) if Y.size else np.empty((0, x_ref.size))
                    x0 = x_ref

            pool[st]["Ys"].append(Y)
            pool[st]["sample_nums"].add(str(sample_num))

        for st in typelist:
            if st not in pool:
                continue
            d = pool[st]
            x = d["x"]
            Yall = np.vstack([Y for Y in d["Ys"] if Y.size]) if d["Ys"] else np.empty((0, x.size))
            y_mean, y_err, n_specs = _mean_and_err(Yall)
            if y_mean is None:
                continue
            curves.append({
                "subtype": st,
                "sample_num": None,
                "label": f"{st} | {region} | n_specs={n_specs} | n_samples={len(d['sample_nums'])}",
                "x": x,
                "y_mean": y_mean,
                "y_err": y_err,
                "n_spectra": n_specs,
                "sample_nums": sorted(d["sample_nums"]),
            })

        if not curves:
            print(f"[PlotRegionAverageSpectra] No curves to plot (key={region_key}).")
            return None

        plt.figure(figsize=(10, 6))
        for c in curves:
            x, ym, ye = c["x"], c["y_mean"], c["y_err"]
            if show_error and ye is not None:
                plt.fill_between(x, ym - ye, ym + ye, alpha=error_alpha, linewidth=0, zorder=1)
            plt.plot(x, ym, label=c["label"], zorder=2)

        _plot_peak_lines(plt.gca())

        if xlim is not None:
            plt.xlim(xlim)
        if ylim is not None:
            plt.ylim(ylim)

        plt.axhline(0, color="k", lw=1.0, alpha=0.5)
        plt.xlabel("Wavenumber (cm$^{-1}$)")
        plt.ylabel("Intensity (a.u.)")
        plt.title(f"{reg_short} region-average spectra ({region}) | AveragebyType={AveragebyType} | err={err_mode.upper()}")
        plt.legend(fontsize="x-small")
        plt.tight_layout()
        plt.show()

        return curves

    # -----------------------------
    # AveragebyType = False
    # -----------------------------
    for sample_num, sample_data in data_dict.items():
        st = _get_subtype(sample_data)
        if st not in typelist_set:
            continue

        sd = _extract_region_spectra(sample_data)
        if sd is None:
            if verbose:
                print(f"[PlotRegionAverageSpectra] Missing {region_key} for sample {sample_num} ({st})")
            continue

        x, Y = _align_stack(sd)
        y_mean, y_err, n_specs = _mean_and_err(Y)
        if y_mean is None:
            continue

        sname = sample_data.get("SampleName", "")
        curves.append({
            "subtype": st,
            "sample_num": str(sample_num),
            "label": f"{st} | {sample_num} | {sname} | {region} | n_specs={n_specs}",
            "x": x,
            "y_mean": y_mean,
            "y_err": y_err,
            "n_spectra": n_specs,
            "sample_nums": [str(sample_num)],
        })

    if not curves:
        print(f"[PlotRegionAverageSpectra] No curves to plot (key={region_key}).")
        return None

    # subtype colours for top-left overview
    subtype_colour_map = {st: f"C{i}" for i, st in enumerate(typelist)}

    # per-sample colours within each subtype panel
    sample_colour_map = {}
    for st in typelist:
        sample_curves = [c for c in curves if c["subtype"] == st]
        cmap = plt.cm.get_cmap("tab10", max(len(sample_curves), 1))
        sample_colour_map[st] = {
            c["sample_num"]: cmap(i) for i, c in enumerate(sample_curves)
        }

    fig, axes = plt.subplots(2, 2, figsize=(14, 10), sharex=True, sharey=True)
    ax_all = axes[0, 0]

    panel_map = {}
    remaining_axes = [axes[0, 1], axes[1, 0], axes[1, 1]]
    for ax, st in zip(remaining_axes, typelist):
        panel_map[st] = ax

    # top-left: all samples, coloured by subtype
    added_subtype_labels = set()
    for c in curves:
        st = c["subtype"]
        x, ym, ye = c["x"], c["y_mean"], c["y_err"]
        colour = subtype_colour_map.get(st, None)

        if show_error and ye is not None:
            ax_all.fill_between(x, ym - ye, ym + ye, alpha=0.15, linewidth=0, color=colour, zorder=1)

        ax_all.plot(
            x, ym,
            color=colour,
            label=st if st not in added_subtype_labels else None,
            zorder=2
        )
        added_subtype_labels.add(st)

    _plot_peak_lines(ax_all)
    ax_all.axhline(0, color="k", lw=1.0, alpha=0.5)
    ax_all.set_title("All subtypes, all samples")
    ax_all.set_ylabel("Intensity (a.u.)")
    ax_all.legend(fontsize="small")

    # individual subtype panels
    for st in typelist:
        ax = panel_map.get(st, None)
        if ax is None:
            continue

        subtype_curves = [c for c in curves if c["subtype"] == st]

        for c in subtype_curves:
            x, ym, ye = c["x"], c["y_mean"], c["y_err"]
            sample_num = c["sample_num"]
            colour = sample_colour_map[st][sample_num]

            if show_error and ye is not None:
                ax.fill_between(x, ym - ye, ym + ye, alpha=0.20, linewidth=0, color=colour, zorder=1)

            ax.plot(
                x, ym,
                color=colour,
                label=sample_num,
                zorder=2
            )

        _plot_peak_lines(ax)
        ax.axhline(0, color="k", lw=1.0, alpha=0.5)
        ax.set_title(f"{st} samples")
        ax.legend(title="Sample", fontsize="x-small", title_fontsize="x-small")

    # global formatting
    for ax in axes[1, :]:
        ax.set_xlabel("Wavenumber (cm$^{-1}$)")
    for ax in axes[:, 0]:
        ax.set_ylabel("Intensity (a.u.)")

    if xlim is not None:
        for ax in axes.flat:
            ax.set_xlim(xlim)
    if ylim is not None:
        for ax in axes.flat:
            ax.set_ylim(ylim)

    fig.suptitle(f"{reg_short} region-average spectra ({region}) | AveragebyType=False | err={err_mode.upper()}", y=0.98)
    plt.tight_layout()
    plt.show()

    return curves



def ExportWeightedMoments_ByRegion(
    data_dict,
    TypestoPlot,
    region="dermis",
    peak_regions=None,
    out_xlsx="weighted_moments_by_region.xlsx",
    use_FP=True,
    use_treated=True,
    verbose=True,
):
    """
    Export weighted moments calculated from the SAMPLE-AVERAGED spectrum
    for each sample/region, not from pointwise moments averaged afterwards.

    Output workbook:
      - Sheet "Averaged": one row per sample
      - One sheet per sample: the averaged spectrum plus metrics for each peak region
      - Sheet "Pooled": pooled by subtype from sample-level values
    """

    if peak_regions is None:
        peak_regions = [
            ("AmideI_1550_1800", (1550, 1750)),
            ("AmideIII_1300_1600", (1400, 1550)),
            ("CH2CH3_1150_1450", (1200, 1400)),
        ]

    typelist = set([str(t).strip() for t in (TypestoPlot if isinstance(TypestoPlot, (list, tuple, set)) else [TypestoPlot])])
    region = str(region).strip().lower()
    reg_short = "FP" if use_FP else "EXT"
    treated_tag = "Treated" if use_treated else "Treated_Full"
    region_key = f"{reg_short}_Spectra_{treated_tag}_{region.capitalize()}"

    def _sheet_safe(name):
        s = str(name)
        for ch in ["/", "\\", "[", "]", "*", "?", ":"]:
            s = s.replace(ch, "_")
        return s[:31] if len(s) > 31 else s

    def _align_stack(spectra_dict):
        keys = list(spectra_dict.keys())
        sc0 = spectra_dict[keys[0]]
        x0 = np.asarray(sc0.spectral_axis, float)

        Ys = []
        for k in keys:
            sc = spectra_dict[k]
            x = np.asarray(sc.spectral_axis, float)
            y = np.asarray(sc.spectral_data, float)
            if (x.shape != x0.shape) or (not np.allclose(x, x0, rtol=0, atol=1e-9)):
                y = np.interp(x0, x, y)
            Ys.append(y)

        Y = np.vstack(Ys) if Ys else np.empty((0, x0.size))
        return x0, Y

    def _weighted_moments_from_mean_spectrum(x, y, xlim):
        x = np.asarray(x, float)
        y = np.asarray(y, float)

        lo, hi = float(xlim[0]), float(xlim[1])
        m = (x >= lo) & (x <= hi) & np.isfinite(x) & np.isfinite(y)
        if not np.any(m):
            return {
                "m1": np.nan,
                "mu2": np.nan,
                "mu3": np.nan,
                "sigma": np.nan,
                "skewness": np.nan,
                "area_w": np.nan,
                "max_intensity": np.nan,
                "neg_area_frac": np.nan,
                "n_points": 0,
            }

        xr = x[m]
        yr = y[m]

        area_total = float(np.trapezoid(np.abs(yr), xr)) if xr.size > 1 else float("nan")
        area_neg = float(np.trapezoid(np.clip(-yr, 0, None), xr)) if xr.size > 1 else float("nan")
        neg_area_frac = (area_neg / area_total) if (np.isfinite(area_total) and area_total > 0) else float("nan")

        offset = -min(0.0, float(np.nanmin(yr)))
        w = np.clip(yr + offset, 0, None)
        wsum = float(np.nansum(w))

        max_intensity = float(np.nanmax(yr)) if xr.size else float("nan")

        if not (np.isfinite(wsum) and wsum > 0):
            return {
                "m1": np.nan,
                "mu2": np.nan,
                "mu3": np.nan,
                "sigma": np.nan,
                "skewness": np.nan,
                "area_w": np.nan,
                "max_intensity": max_intensity,
                "neg_area_frac": neg_area_frac,
                "n_points": int(xr.size),
            }

        m1 = float(np.nansum(w * xr) / wsum)
        mu2 = float(np.nansum(w * (xr - m1) ** 2) / wsum)
        mu3 = float(np.nansum(w * (xr - m1) ** 3) / wsum)
        sigma = float(np.sqrt(mu2)) if (np.isfinite(mu2) and mu2 >= 0) else float("nan")
        skew = float(mu3 / (mu2 ** 1.5)) if (np.isfinite(mu2) and mu2 > 0 and np.isfinite(mu3)) else float("nan")
        area_w = float(np.trapezoid(w, xr)) if xr.size > 1 else float("nan")

        return {
            "m1": m1,
            "mu2": mu2,
            "mu3": mu3,
            "sigma": sigma,
            "skewness": skew,
            "area_w": area_w,
            "max_intensity": max_intensity,
            "neg_area_frac": neg_area_frac,
            "n_points": int(xr.size),
        }

    rows_avg = []
    per_sample_tables = {}
    matched_any = False

    for sample_num, sample_data in data_dict.items():
        st = str(sample_data.get("Subtype", "")).strip()
        if st not in typelist:
            continue

        sd = sample_data.get(region_key, None)
        if not (isinstance(sd, dict) and len(sd)):
            if verbose:
                print(f"[ExportWeightedMoments_ByRegion] Missing {region_key} for sample {sample_num} ({st})")
            continue

        matched_any = True

        x_mean, Y = _align_stack(sd)
        if Y.size == 0:
            continue
        y_mean = np.nanmean(Y, axis=0)

        base_meta = {
            "SampleNumber": str(sample_num),
            "SampleName": sample_data.get("SampleName", ""),
            "Subtype": st,
            "RegionKey": region_key,
            "Region": region,
            "Wavelength": sample_data.get("Wavelength", ""),
            "Substrate": sample_data.get("Substrate", ""),
            "ScanType": sample_data.get("ScanType", ""),
            "Accumulations": sample_data.get("Accumulations", ""),
            "Exposure": sample_data.get("Exposure", ""),
            "OtherInfo": sample_data.get("OtherInfo", ""),
            "NpointsRegion": int(Y.shape[0]),
        }

        # one-row summary for Averaged
        avg_row = dict(base_meta)

        # per-sample sheet contains averaged spectrum and repeated metrics
        df_points = pd.DataFrame({
            "Wave": x_mean,
            "MeanSpectrum": y_mean,
        })

        for region_name, xlim in peak_regions:
            moms = _weighted_moments_from_mean_spectrum(x_mean, y_mean, xlim)
            for col, val in moms.items():
                avg_row[f"{region_name}_{col}_mean"] = val
                df_points[f"{region_name}_{col}"] = val

        rows_avg.append(avg_row)

        sname = sample_data.get("SampleName", "")
        sheet = _sheet_safe(sname if sname else f"{sample_num}_{st}")
        if sheet in per_sample_tables:
            sheet = _sheet_safe(f"{sheet}_{sample_num}")
        per_sample_tables[sheet] = df_points

        if verbose:
            print(f"\nSample {sample_num} | Subtype {st}")
            for region_name, _ in peak_regions:
                print(
                    f"  {region_name}: "
                    f"m1={avg_row.get(f'{region_name}_m1_mean', np.nan):.6f}, "
                    f"mu2={avg_row.get(f'{region_name}_mu2_mean', np.nan):.6f}, "
                    f"height={avg_row.get(f'{region_name}_max_intensity_mean', np.nan):.6f}"
                )

    if not matched_any:
        print(f"[ExportWeightedMoments_ByRegion] No samples matched TypestoPlot={sorted(typelist)} with key={region_key}.")
        return None, None

    df_avg = pd.DataFrame(rows_avg)

    pooled_rows = []
    metric_cols = [c for c in df_avg.columns if c.endswith("_mean") and c not in base_meta]

    for st, df_st in df_avg.groupby("Subtype", dropna=False):
        row = {
            "Subtype": st,
            "N_samples": int(df_st["SampleNumber"].nunique()) if "SampleNumber" in df_st else np.nan,
            "Region": region,
            "RegionKey": region_key,
            "TypesFilter": ",".join(sorted(typelist)),
        }
        for c in metric_cols:
            vals = df_st[c]
            row[f"{c}_pooled_mean"] = float(np.nanmean(vals)) if len(vals) else np.nan
            row[f"{c}_pooled_std"] = float(np.nanstd(vals, ddof=1)) if np.sum(np.isfinite(vals)) > 1 else np.nan
        pooled_rows.append(row)

    df_pooled = pd.DataFrame(pooled_rows)

    out_xlsx = Path(out_xlsx)
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        df_avg.to_excel(writer, sheet_name="Averaged", index=False)
        if df_pooled is not None and not df_pooled.empty:
            df_pooled.to_excel(writer, sheet_name="Pooled", index=False)
        for sheet, df in per_sample_tables.items():
            df.to_excel(writer, sheet_name=sheet, index=False)

    if verbose:
        print(f"[ExportWeightedMoments_ByRegion] Wrote: {out_xlsx}")
        print(f"  - Averaged rows: {len(df_avg)}")
        print(f"  - Per-sample sheets: {len(per_sample_tables)}")
        print(f"  - Region used: {region_key}")
        print(f"  - Types: {sorted(typelist)}")

    return df_avg, per_sample_tables

def BinRegionAndPlot(
    data_dict,
    TypestoPlot,
    NBins=5,
    region="dermis",
    use_FP=True,
    PlotType="ByType",   # "ByType" or "ByBin"
    colours=None,
    linestyles=None,
    show_error=True,
    error_alpha=0.3,
    peak_regions=None,
):

    region = region.lower()
    reg_short = "FP" if use_FP else "EXT"
    region_key = f"{reg_short}_Spectra_Treated_{region.capitalize()}"

    typelist = list(TypestoPlot if isinstance(TypestoPlot, (list, tuple, set)) else [TypestoPlot])

    # --------------------------------------------
    # Step 1: bin each sample individually
    # --------------------------------------------
    subtype_pool = {st: [[] for _ in range(NBins)] for st in typelist}

    for sample_num, sample_data in data_dict.items():
        st = sample_data.get("Subtype")
        if st not in typelist:
            continue

        sd = sample_data.get(region_key)
        if not isinstance(sd, dict) or not sd:
            continue

        keys_order = list(sd.keys())

        # apply direction
        direc = sample_data.get("direction", "").strip().lower()
        if direc == "back":
            keys_order = list(reversed(keys_order))

        n = len(keys_order)
        if n == 0:
            continue

        # normalised positions 0→1
        norm_pos = np.linspace(0, 1, n)

        # bin edges
        edges = np.linspace(0, 1, NBins + 1)

        ref_axis = np.asarray(sd[keys_order[0]].spectral_axis, float)

        for b in range(NBins):
            if b == NBins - 1:
                mask = (norm_pos >= edges[b]) & (norm_pos <= edges[b + 1])
            else:
                mask = (norm_pos >= edges[b]) & (norm_pos < edges[b + 1])

            idxs = np.where(mask)[0]
            if idxs.size == 0:
                continue

            Ys = []
            for i in idxs:
                sc = sd[keys_order[i]]
                x = np.asarray(sc.spectral_axis, float)
                y = np.asarray(sc.spectral_data, float)
                if not np.allclose(x, ref_axis, rtol=0, atol=1e-9):
                    y = np.interp(ref_axis, x, y)
                Ys.append(y)

            Ybin = np.mean(np.vstack(Ys), axis=0)
            subtype_pool[st][b].append(Ybin)

    # --------------------------------------------
    # Step 2: merge across samples (per subtype)
    # --------------------------------------------
    binned_avg = {}

    for st in typelist:
        bins = []
        for b in range(NBins):
            if not subtype_pool[st][b]:
                bins.append(None)
                continue
            stack = np.vstack(subtype_pool[st][b])
            mean = np.mean(stack, axis=0)
            std = np.std(stack, axis=0, ddof=1) if stack.shape[0] > 1 else np.zeros_like(mean)
            sem = std / np.sqrt(stack.shape[0]) if stack.shape[0] > 1 else std
            bins.append({
                "Wave": ref_axis,
                "Mean": mean,
                "SEM": sem,
                "N": stack.shape[0]
            })
        binned_avg[st] = bins

    # --------------------------------------------
    # global y-limits across all plots
    # --------------------------------------------
    ymins, ymaxs = [], []
    for st in typelist:
        for bin_data in binned_avg.get(st, []):
            if bin_data is None:
                continue
            y = bin_data["Mean"]
            err = bin_data["SEM"]
            if show_error and err is not None:
                ymins.append(np.nanmin(y - err))
                ymaxs.append(np.nanmax(y + err))
            else:
                ymins.append(np.nanmin(y))
                ymaxs.append(np.nanmax(y))

    global_ylim = None
    if ymins and ymaxs:
        ymin = float(np.nanmin(ymins))
        ymax = float(np.nanmax(ymaxs))
        if np.isfinite(ymin) and np.isfinite(ymax):
            pad = 0.05 * (ymax - ymin) if ymax > ymin else 0.05 * max(abs(ymax), 1.0)
            global_ylim = (ymin - pad, ymax + pad)

    # --------------------------------------------
    # helper for plotting peak-region boundaries
    # --------------------------------------------
    def _plot_peak_lines():
        if not peak_regions:
            return
        region_styles = ["--", ":", "-."]
        for i, (_name, (x1, x2)) in enumerate(peak_regions):
            ls = region_styles[i % len(region_styles)]
            plt.axvline(x1, color="k", linestyle=ls, alpha=0.4, linewidth=1,
                        label="Peak region bounds" if i == 0 else None)
            plt.axvline(x2, color="k", linestyle=ls, alpha=0.4, linewidth=1)

    # --------------------------------------------
    # Step 3: Plotting
    # --------------------------------------------
    PlotType = str(PlotType).strip()

    if PlotType == "ByType":
        # One plot per subtype, legend = bin number, coloured with jet
        cmap = plt.cm.get_cmap("jet", NBins)

        for st in typelist:
            plt.figure(figsize=(10, 6))

            bins = binned_avg.get(st, [])
            for b, bin_data in enumerate(bins):
                if bin_data is None:
                    continue

                x = bin_data["Wave"]
                y = bin_data["Mean"]
                err = bin_data["SEM"]
                colour = cmap(b)

                plt.plot(x, y, label=f"Bin {b+1}", color=colour)

                if show_error:
                    plt.fill_between(
                        x, y - err, y + err,
                        color=colour,
                        alpha=error_alpha,
                        linewidth=0
                    )

            _plot_peak_lines()

            if global_ylim is not None:
                plt.ylim(global_ylim)

            plt.xlabel("Wavenumber (cm$^{-1}$)")
            plt.ylabel("Intensity (a.u.)")
            plt.title(f"{st} — {region.capitalize()} — NBins={NBins}")
            plt.legend(fontsize="x-small")
            plt.tight_layout()
            plt.show()

    elif PlotType == "ByBin":
        # One plot per bin, legend = subtype
        for b in range(NBins):
            plt.figure(figsize=(10, 6))

            for st in typelist:
                bin_data = binned_avg.get(st, [None] * NBins)[b]
                if bin_data is None:
                    continue

                x = bin_data["Wave"]
                y = bin_data["Mean"]
                err = bin_data["SEM"]

                colour = colours.get(st) if colours else None
                linestyle = linestyles.get(st) if linestyles else "-"

                plt.plot(
                    x, y,
                    label=f"{st}",
                    color=colour,
                    linestyle=linestyle
                )

                if show_error:
                    plt.fill_between(
                        x, y - err, y + err,
                        color=colour,
                        alpha=error_alpha,
                        linewidth=0
                    )

            _plot_peak_lines()

            if global_ylim is not None:
                plt.ylim(global_ylim)

            plt.xlabel("Wavenumber (cm$^{-1}$)")
            plt.ylabel("Intensity (a.u.)")
            plt.title(f"Bin {b+1} — {region.capitalize()} — NBins={NBins}")
            plt.legend(fontsize="x-small")
            plt.tight_layout()
            plt.show()

    else:
        print(f"[BinRegionAndPlot] Unknown PlotType='{PlotType}' (use 'ByType' or 'ByBin').")

    return binned_avg

def AppendBinsToWeightedMoments(NBins, xlsx_path):
    """
    Rebuild the 'Bins' sheet from sample-averaged spectra stored in each sample sheet.

    Each sample sheet is expected to contain:
      - Wave
      - MeanSpectrum
      - repeated constant metric columns such as:
            <PeakRegion>_m1
            <PeakRegion>_mu2
            <PeakRegion>_sigma
            <PeakRegion>_max_intensity
      - or metadata columns may be absent

    This function does NOT use pointwise rows anymore.
    It bins the averaged spectrum itself into NBins over the spectrum index only if needed.

    Since the current export is one averaged spectrum per sample, the bin summaries are
    built from the already-exported metric columns and pooled by subtype.
    """

    verbose = True
    xlsx_path = Path(xlsx_path)

    if not xlsx_path.exists():
        print(f"[AppendBinsToWeightedMoments] File not found: {xlsx_path}")
        return None

    xl = pd.ExcelFile(xlsx_path)
    sheet_names = xl.sheet_names

    ignore = {"Averaged", "Pooled", "Bins", "BinStats", "BinStats_Summary"}
    sample_sheets = [s for s in sheet_names if s not in ignore]

    if verbose:
        print(f"[AppendBinsToWeightedMoments] Found {len(sample_sheets)} sample sheets")

    # subtype lookup comes from Averaged sheet now
    df_avg = xl.parse("Averaged")
    if "SampleNumber" not in df_avg.columns or "Subtype" not in df_avg.columns:
        print("[AppendBinsToWeightedMoments] 'Averaged' sheet missing SampleNumber/Subtype.")
        return None

    # build lookup by sample number and sample name
    subtype_lookup = {}
    for _, row in df_avg.iterrows():
        sn = str(row.get("SampleNumber", "")).strip()
        st = str(row.get("Subtype", "")).strip()
        sname = str(row.get("SampleName", "")).strip()
        if sn:
            subtype_lookup[sn] = st
        if sname:
            subtype_lookup[sname] = st

    rows_out = []

    for sheet in sample_sheets:
        df = xl.parse(sheet)
        if df.empty:
            continue

        # infer subtype from Averaged mapping
        sheet_name = str(sheet).strip()
        subtype = subtype_lookup.get(sheet_name, None)
        
        # try sample number parsed from sheet name like "356_D14"
        if subtype is None:
            sn_from_sheet = sheet_name.split("_")[0].strip()
            subtype = subtype_lookup.get(sn_from_sheet, None)
        
        # if still missing, try SampleNumber column if present
        if subtype is None and "SampleNumber" in df.columns:
            nonnull = df["SampleNumber"].dropna()
            if not nonnull.empty:
                sn0 = str(nonnull.iloc[0]).strip()
                subtype = subtype_lookup.get(sn0, None)

        if subtype is None:
            # final fallback: skip if we cannot map
            if verbose:
                print(f"[AppendBinsToWeightedMoments] Could not infer subtype for sheet '{sheet}', skipping.")
            continue

        metric_cols = [
            c for c in df.columns
            if any(c.endswith(suf) for suf in (
                "_m1", "_mu2", "_mu3", "_sigma",
                "_skewness", "_area_w", "_max_intensity", "_neg_area_frac"
            ))
        ]

        if not metric_cols:
            continue

        # metrics are constant down the per-sample sheet now, so just take first row
        row0 = df.iloc[0]

        # because there is no spatial pointwise table anymore, one sample contributes
        # the same overall value to each "bin" placeholder. This preserves downstream plotting.
        for b in range(1, NBins + 1):
            row = {
                "Subtype": subtype,
                "SampleSheet": sheet,
                "Bin": b,
                "Npoints_in_bin": 1
            }

            for col in metric_cols:
                val = float(row0[col]) if pd.notna(row0[col]) else np.nan
                row[f"{col}_mean"] = val
                row[f"{col}_std"] = 0.0

            rows_out.append(row)

    df_bins_samples = pd.DataFrame(rows_out)

    if df_bins_samples.empty:
        print("[AppendBinsToWeightedMoments] No usable sample-sheet data found.")
        return None

    pooled_rows = []
    group_cols = ["Subtype", "Bin"]
    metric_mean_cols = [c for c in df_bins_samples.columns if c.endswith("_mean")]

    for (st, b), g in df_bins_samples.groupby(group_cols):
        row = {
            "Subtype": st,
            "Bin": b,
            "Nsamples_in_bin": len(g)
        }

        for col in metric_mean_cols:
            vals = g[col]
            row[f"{col}_pooled_mean"] = float(np.nanmean(vals)) if len(vals) else np.nan
            row[f"{col}_pooled_std"] = float(np.nanstd(vals, ddof=1)) if np.sum(np.isfinite(vals)) > 1 else np.nan

        pooled_rows.append(row)

    df_bins_pooled = pd.DataFrame(pooled_rows)

    with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df_bins_pooled.to_excel(writer, sheet_name="Bins", index=False)

    if verbose:
        print(f"[AppendBinsToWeightedMoments] Added 'Bins' sheet to {xlsx_path}")
        print(f"  Rows written: {len(df_bins_pooled)}")

    return df_bins_pooled

def PlotBinsWeightedMoment(
    xlsx_path,
    Peak="AmideI_LEFT_1550_1800",
    WM="1",
    colours=None,
    linestyles=None,
    figsize=(8,6),
    show_error=True,
    error_alpha=0.25,
    verbose=True
):
    """
    Plot weighted moment across bins for each subtype.

    Peak should ideally be the full peak-region name, e.g.
    'AmideI_LEFT_1550_1800', 'AmideI_MIDDLE_1550_1800', etc.

    WM:
        '1', '2', '3', 'sigma', 'skew', 'area', 'neg'
    """

    xlsx_path = Path(xlsx_path)
    df = pd.read_excel(xlsx_path, sheet_name="Bins")

    if df.empty:
        print("[PlotBinsWeightedMoment] Bins sheet empty.")
        return None

    if isinstance(WM, (list, tuple)) and len(WM) == 1:
        WM = WM[0]
    if isinstance(Peak, (list, tuple)) and len(Peak) == 1:
        Peak = Peak[0]

    WM = str(WM).strip().lower()
    Peak = str(Peak).strip()

    wm_map = {
        "1": "_m1",
        "2": "_mu2",
        "3": "_mu3",
        "sigma": "_sigma",
        "skew": "_skewness",
        "area": "_area_w",
        "neg": "_neg_area_frac"
    }

    if WM not in wm_map:
        print(f"[PlotBinsWeightedMoment] Unknown WM='{WM}'")
        return None

    suffix = wm_map[WM]

    pooled_mean_cols = [c for c in df.columns if c.endswith("_pooled_mean")]

    # Prefer exact region match
    exact_cols = [
        c for c in pooled_mean_cols
        if c.startswith(f"{Peak}{suffix}") and c.endswith("_pooled_mean")
    ]

    # Fallback: partial match
    partial_cols = [
        c for c in pooled_mean_cols
        if (Peak.lower() in c.lower()) and suffix in c
    ]

    candidate_cols = exact_cols if exact_cols else partial_cols

    if not candidate_cols:
        print(f"[PlotBinsWeightedMoment] No column found for Peak={Peak}, WM={WM}")
        return None

    if len(candidate_cols) > 1 and verbose:
        print(f"[PlotBinsWeightedMoment] Multiple matches found for Peak={Peak}, WM={WM}:")
        for c in candidate_cols:
            print(f"  - {c}")
        print(f"[PlotBinsWeightedMoment] Using: {candidate_cols[0]}")

    col_mean = candidate_cols[0]
    col_std = col_mean.replace("_pooled_mean", "_pooled_std")

    if verbose:
        print(f"[PlotBinsWeightedMoment] Using column: {col_mean}")

    plt.figure(figsize=figsize)

    subtypes = sorted(df["Subtype"].unique())

    for st in subtypes:
        dsub = df[df["Subtype"] == st].sort_values("Bin")

        x = dsub["Bin"].values
        y = dsub[col_mean].values
        yerr = dsub[col_std].values if col_std in dsub.columns else None

        colour = colours.get(st) if colours else None
        linestyle = linestyles.get(st) if linestyles else "-"

        if show_error and yerr is not None:
            plt.errorbar(
                x,
                y,
                yerr=yerr,
                marker="o",
                linestyle=linestyle,
                color=colour,
                capsize=4,
                elinewidth=1.2,
                linewidth=2,
                label=st
            )
        else:
            plt.plot(
                x,
                y,
                marker="o",
                label=st,
                color=colour,
                linestyle=linestyle
            )

    plt.xlabel("Normalised Bin Position")
    plt.ylabel(f"{Peak} – Weighted Moment {WM}")
    plt.title(f"{Peak} | Weighted Moment {WM} across dermis bins")
    plt.legend()
    plt.tight_layout()
    plt.show()

    return col_mean



def BinStats_TTests_and_MixedLM_Summary(
    xlsx_path,
    TypestoPlot=("CT", "D7", "D14"),
    sheet_out="BinStats",
    sheet_summary="BinStats_Summary",
    alpha_levels=(0.05, 0.01, 0.001),
    verbose=True,
):
    """
    Sample-level stats only, using the 'Averaged' sheet.

    Since metrics are now calculated from the sample-averaged spectrum,
    there is no longer a point-level table suitable for MixedLM.

    This function now:
      1) reads the Averaged sheet
      2) runs Welch t-tests between subtypes for each metric column
      3) writes detailed and summary significance sheets

    Returns
    -------
    df_t : detailed Welch t-test results
    df_m : empty DataFrame (kept for backwards compatibility)
    """
    import numpy as np
    import pandas as pd
    import warnings

    from pathlib import Path
    from scipy.stats import ttest_ind
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    xlsx_path = Path(xlsx_path)

    def _stars(p):
        if p is None or (isinstance(p, float) and not np.isfinite(p)):
            return "NA"
        a05, a01, a001 = alpha_levels
        if p < a001:
            return "***"
        if p < a01:
            return "**"
        if p < a05:
            return "*"
        return "ns"

    def _pairwise_groups(groups):
        gs = list(groups)
        out = []
        for i in range(len(gs)):
            for j in range(i + 1, len(gs)):
                out.append((gs[i], gs[j]))
        return out

    def _parse_metric(col):
        suffixes = (
            "_m1_mean",
            "_mu2_mean",
            "_mu3_mean",
            "_sigma_mean",
            "_skewness_mean",
            "_area_w_mean",
            "_max_intensity_mean",
            "_neg_area_frac_mean",
        )
        for suf in suffixes:
            if col.endswith(suf):
                peak = col[:-len(suf)]
                wm = suf.replace("_mean", "").lstrip("_")
                return peak, wm
        return None, None

    def _welch_ttests(df_avg, atol=1e-12):
        rows = []

        metric_cols = []
        for c in df_avg.columns:
            peak, wm = _parse_metric(c)
            if peak is not None:
                metric_cols.append((c, peak, wm))

        for col, peak, wm in metric_cols:
            dmet = df_avg[["SampleNumber", "Subtype", col]].copy()
            dmet = dmet.replace([np.inf, -np.inf], np.nan).dropna(subset=[col])

            for g1, g2 in _pairwise_groups(TypestoPlot):
                x1 = dmet.loc[dmet["Subtype"] == g1, col].astype(float).dropna().values
                x2 = dmet.loc[dmet["Subtype"] == g2, col].astype(float).dropna().values

                row = {
                    "Test": "Welch_ttest_perSampleValue",
                    "Peak": peak,
                    "WM": wm,
                    "Metric": col,
                    "Group1": g1,
                    "Group2": g2,
                    "N1": len(x1),
                    "N2": len(x2),
                    "Mean1": float(np.mean(x1)) if len(x1) else np.nan,
                    "Mean2": float(np.mean(x2)) if len(x2) else np.nan,
                    "Diff(1-2)": float(np.mean(x1) - np.mean(x2)) if (len(x1) and len(x2)) else np.nan,
                    "t": np.nan,
                    "p": np.nan,
                    "Status": "ok",
                }

                if len(x1) < 2 or len(x2) < 2:
                    row["Status"] = "too_few_samples"
                    rows.append(row)
                    continue

                v1 = np.nanvar(x1, ddof=1)
                v2 = np.nanvar(x2, ddof=1)

                if (not np.isfinite(v1) or v1 <= atol) and (not np.isfinite(v2) or v2 <= atol):
                    if np.isclose(row["Mean1"], row["Mean2"], atol=atol, rtol=0):
                        row["t"] = 0.0
                        row["p"] = 1.0
                        row["Status"] = "both_constant_equal"
                    else:
                        row["Status"] = "both_constant_different"
                    rows.append(row)
                    continue

                with warnings.catch_warnings():
                    warnings.filterwarnings(
                        "ignore",
                        message="Precision loss occurred in moment calculation due to catastrophic cancellation.*",
                        category=RuntimeWarning,
                    )
                    t, p = ttest_ind(x1, x2, equal_var=False, nan_policy="omit")

                row["t"] = float(t) if np.isfinite(t) else np.nan
                row["p"] = float(p) if np.isfinite(p) else np.nan
                rows.append(row)

        return pd.DataFrame(rows)

    def _make_star_tables(df_detail):
        out = {}
        if df_detail.empty:
            return out

        peaks = sorted(df_detail["Peak"].dropna().unique())
        for pk in peaks:
            dpk = df_detail[df_detail["Peak"] == pk].copy()
            wms = sorted(dpk["WM"].dropna().unique())

            comps = sorted({f"{a} vs {b}" for a, b in zip(dpk["Group1"], dpk["Group2"])})
            tab = pd.DataFrame(index=comps, columns=wms, data="NA")

            for _, r in dpk.iterrows():
                comp = f"{r['Group1']} vs {r['Group2']}"
                wm = r["WM"]
                if comp in tab.index and wm in tab.columns:
                    tab.loc[comp, wm] = _stars(r.get("p"))

            tab.index.name = "Comparison"
            tab.columns.name = "WM"
            out[pk] = tab

        return out

    def _write_blocks(ws, title, df, start_row):
        ws.cell(row=start_row, column=1, value=title)
        r = start_row + 1
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)
            r += 1
        return r + 2

    def _write_star_tables(ws, title, tables_dict, start_row):
        ws.cell(row=start_row, column=1, value=title)
        r = start_row + 1
        for pk, tab in tables_dict.items():
            ws.cell(row=r, column=1, value=str(pk))
            r += 1
            tab2 = tab.copy()
            tab2.insert(0, tab2.index.name or "Comparison", tab2.index)
            tab2 = tab2.reset_index(drop=True)
            for row in dataframe_to_rows(tab2, index=False, header=True):
                ws.append(row)
                r += 1
            r += 2
        return r

    # -------------------- read Averaged sheet --------------------
    df_avg = pd.read_excel(xlsx_path, sheet_name="Averaged")
    if df_avg.empty:
        raise ValueError("Averaged sheet is empty.")

    df_avg = df_avg[df_avg["Subtype"].astype(str).isin(TypestoPlot)].copy()
    if df_avg.empty:
        raise ValueError("No matching subtypes found in Averaged sheet.")

    if verbose:
        print(f"[BinStats] samples used from Averaged sheet: {len(df_avg)}")

    # -------------------- run stats --------------------
    df_t = _welch_ttests(df_avg)
    df_m = pd.DataFrame(columns=["Test", "Note"])
    if not df_t.empty:
        df_t["sig"] = df_t["p"].apply(_stars)

    t_tables = _make_star_tables(df_t)

    # -------------------- write workbook sheets --------------------
    wb = load_workbook(xlsx_path)

    if sheet_out in wb.sheetnames:
        del wb[sheet_out]
    ws = wb.create_sheet(sheet_out)

    row = 1
    row = _write_blocks(ws, "Welch t-tests between groups (sample-level values from Averaged sheet)", df_t, row)

    if sheet_summary in wb.sheetnames:
        del wb[sheet_summary]
    ws2 = wb.create_sheet(sheet_summary)

    r2 = 1
    r2 = _write_star_tables(ws2, "T-test significance tables", t_tables, r2)

    wb.save(xlsx_path)

    if verbose:
        print(f"[BinStats] Wrote sheets: '{sheet_out}', '{sheet_summary}' -> {xlsx_path}")

    return df_t, df_m



def PlotOverallWeightedMomentBars(
    xlsx_path,
    TypestoPlot=("CT", "D7", "D14"),
    WMs=("1", "2"),
    scale_mode="auto",   # "auto", "raw", "delta_from_control", "minmax", "first_group"
    colours=None,
    figsize=(12, 6),
    error_mode="sem",    # "sem" or "std"
    verbose=True
):
    """
    Plot one overall value per peak region for WM1 / WM2 / height.

    Reads the 'Averaged' sheet where each row is one sample and columns contain
    sample-level metrics calculated from the sample-averaged spectrum.

    scale_mode:
      - "raw": use raw values directly
      - "delta_from_control": subtract CT mean within each peak region
      - "minmax": scale each peak region to 0-1
      - "first_group": divide by first group mean within each peak region
      - "auto": use raw for height/intensity, delta_from_control for WM1/WM2
    """

    xlsx_path = Path(xlsx_path)
    df = pd.read_excel(xlsx_path, sheet_name="Averaged")

    if df.empty:
        print("[PlotOverallWeightedMomentBars] Averaged sheet empty.")
        return None

    typelist = [str(t).strip() for t in TypestoPlot]
    df = df[df["Subtype"].astype(str).isin(typelist)].copy()

    if df.empty:
        print("[PlotOverallWeightedMomentBars] No matching subtypes found.")
        return None

    wm_map = {
        "1": "_m1_mean",
        "2": "_mu2_mean",
        "3": "_mu3_mean",
        "sigma": "_sigma_mean",
        "skew": "_skewness_mean",
        "area": "_area_w_mean",
        "height": "_max_intensity_mean",
        "intensity": "_max_intensity_mean",
        "neg": "_neg_area_frac_mean"
    }

    label_map = {
        "1": "WM1 shift from CT (cm$^{-1}$)",
        "2": "WM2 shift from CT",
        "3": "WM3 shift from CT",
        "sigma": "Sigma shift from CT",
        "skew": "Skewness shift from CT",
        "area": "Area change from CT",
        "height": "Peak height (a.u.)",
        "intensity": "Peak height (a.u.)",
        "neg": "Negative area fraction change from CT",
    }

    title_map = {
        "1": "WM1 relative to control",
        "2": "WM2 relative to control",
        "3": "WM3 relative to control",
        "sigma": "Sigma relative to control",
        "skew": "Skewness relative to control",
        "area": "Area relative to control",
        "height": "Peak height",
        "intensity": "Peak height",
        "neg": "Negative area fraction relative to control",
    }

    def _peak_name_from_col(col, suffix):
        if col.endswith(suffix):
            return col[:-len(suffix)]
        return col

    def _transform_group(g, value_col, mode):
        g = g.copy()
        vals = g[value_col].astype(float).values

        if mode == "raw":
            g["ValuePlot"] = vals
            return g

        if mode == "minmax":
            vmin = np.nanmin(vals)
            vmax = np.nanmax(vals)
            if np.isfinite(vmin) and np.isfinite(vmax) and vmax > vmin:
                g["ValuePlot"] = (vals - vmin) / (vmax - vmin)
            else:
                g["ValuePlot"] = 0.0
            return g

        if mode == "first_group":
            base_group = typelist[0]
            base = g.loc[g["Subtype"] == base_group, value_col].astype(float)
            denom = np.nanmean(base) if len(base) else np.nan
            if np.isfinite(denom) and denom != 0:
                g["ValuePlot"] = vals / denom
            else:
                g["ValuePlot"] = np.nan
            return g

        if mode == "delta_from_control":
            base_group = typelist[0]
            base = g.loc[g["Subtype"] == base_group, value_col].astype(float)
            base_mean = np.nanmean(base) if len(base) else np.nan
            if np.isfinite(base_mean):
                g["ValuePlot"] = vals - base_mean
            else:
                g["ValuePlot"] = np.nan
            return g

        raise ValueError(f"Unknown scale_mode='{mode}'")

    outputs = {}

    for WM in WMs:
        WM = str(WM).strip().lower()
        if WM not in wm_map:
            print(f"[PlotOverallWeightedMomentBars] Unknown WM='{WM}'")
            continue

        suffix = wm_map[WM]

        metric_cols = [c for c in df.columns if c.endswith(suffix)]
        if not metric_cols:
            print(f"[PlotOverallWeightedMomentBars] No columns found for WM={WM}")
            continue

        if verbose:
            print(f"[PlotOverallWeightedMomentBars] WM={WM} columns:")
            for c in metric_cols:
                print(f"  - {c}")

        long_rows = []
        for _, row in df.iterrows():
            for col in metric_cols:
                long_rows.append({
                    "SampleNumber": row["SampleNumber"],
                    "Subtype": str(row["Subtype"]).strip(),
                    "PeakRegion": _peak_name_from_col(col, suffix),
                    "Value": row[col]
                })

        dlong = pd.DataFrame(long_rows)
        dlong = dlong.replace([np.inf, -np.inf], np.nan).dropna(subset=["Value"])

        if dlong.empty:
            print(f"[PlotOverallWeightedMomentBars] No usable data for WM={WM}")
            continue

        if scale_mode == "auto":
            mode = "raw" if WM in ("height", "intensity") else "delta_from_control"
        else:
            mode = scale_mode

        transformed = []
        for peak, g in dlong.groupby("PeakRegion", sort=False):
            transformed.append(_transform_group(g.copy(), "Value", mode))

        dlong = pd.concat(transformed, ignore_index=True)

        summary_rows = []
        for (peak, st), g in dlong.groupby(["PeakRegion", "Subtype"]):
            vals = g["ValuePlot"].astype(float).values
            mean = np.nanmean(vals) if len(vals) else np.nan
            if error_mode == "std":
                err = np.nanstd(vals, ddof=1) if np.sum(np.isfinite(vals)) > 1 else 0.0
            else:
                err = (
                    np.nanstd(vals, ddof=1) / np.sqrt(np.sum(np.isfinite(vals)))
                    if np.sum(np.isfinite(vals)) > 1 else 0.0
                )

            summary_rows.append({
                "PeakRegion": peak,
                "Subtype": st,
                "Mean": mean,
                "Err": err,
                "N": np.sum(np.isfinite(vals))
            })

        dsum = pd.DataFrame(summary_rows)

        peak_order = list(dict.fromkeys(dlong["PeakRegion"].tolist()))
        x = np.arange(len(peak_order))
        n_groups = len(typelist)
        width = 0.8 / max(n_groups, 1)

        plt.figure(figsize=figsize)

        for i, st in enumerate(typelist):
            dsub = (
                dsum[dsum["Subtype"] == st]
                .set_index("PeakRegion")
                .reindex(peak_order)
                .reset_index()
            )

            xpos = x - 0.4 + width/2 + i * width
            y = dsub["Mean"].values
            yerr = dsub["Err"].values

            colour = colours.get(st) if colours else None

            plt.bar(
                xpos,
                y,
                width=width,
                label=st,
                color=colour,
                alpha=0.9
            )

            plt.errorbar(
                xpos,
                y,
                yerr=yerr,
                fmt="none",
                ecolor="k",
                elinewidth=1,
                capsize=3
            )

        if mode == "delta_from_control":
            plt.axhline(0, color="k", lw=1.0, alpha=0.5)

        plt.xticks(x, peak_order, rotation=45, ha="right")
        plt.ylabel(label_map.get(WM, WM))
        plt.title(f"Overall peak-region comparison | {title_map.get(WM, WM)}")
        plt.legend()
        plt.tight_layout()
        plt.show()

        outputs[WM] = {
            "long": dlong,
            "summary": dsum,
            "mode": mode
        }

    return outputs

def PlotPeakCentreAndSpread(
    xlsx_path,
    peak_regions,
    TypestoPlot=("CT", "D7", "D14"),
    colours=None,
    linestyles=None,
    figsize=(12, 7),
    normalise_x=False,
    x_mode="absolute",   # "absolute" or "delta_from_first"
    show_region_bounds=True,
    verbose=True
):
    """
    Plot WM1 (centre) and WM2-derived spread together.

    For each peak region:
      - point = mean WM1 across samples in subtype
      - horizontal bar = WM1 ± sigma, where sigma = sqrt(WM2)

    Parameters
    ----------
    xlsx_path : str or Path
        Workbook containing the 'Averaged' sheet.
    peak_regions : list of tuples
        e.g. [("AmideI_LEFT_1530_1590", (1530,1590)), ...]
        Used for order and optional background spans.
    TypestoPlot : iterable
        Subtypes to include, e.g. ("CT","D7","D14")
    x_mode : str
        "absolute" -> x axis is actual WM1 position
        "delta_from_first" -> x axis is WM1 shift relative to first subtype in TypestoPlot
    normalise_x : bool
        If True, normalises WM1 within each peak region to 0–1 before plotting.
        Usually False is better here.
    """

    xlsx_path = Path(xlsx_path)
    df = pd.read_excel(xlsx_path, sheet_name="Averaged")

    if df.empty:
        print("[PlotPeakCentreAndSpread] Averaged sheet empty.")
        return None

    typelist = [str(t).strip() for t in TypestoPlot]
    df = df[df["Subtype"].astype(str).isin(typelist)].copy()

    if df.empty:
        print("[PlotPeakCentreAndSpread] No matching subtypes found.")
        return None

    # build long table: one row per sample × peak region
    rows = []
    for peak_name, (lo, hi) in peak_regions:
        col_m1 = f"{peak_name}_m1_mean"
        col_mu2 = f"{peak_name}_mu2_mean"

        if col_m1 not in df.columns or col_mu2 not in df.columns:
            if verbose:
                print(f"[PlotPeakCentreAndSpread] Missing columns for {peak_name}")
            continue

        for _, r in df.iterrows():
            mu2 = float(r[col_mu2]) if pd.notna(r[col_mu2]) else np.nan
            sigma = np.sqrt(mu2) if np.isfinite(mu2) and mu2 >= 0 else np.nan

            rows.append({
                "SampleNumber": r["SampleNumber"],
                "Subtype": str(r["Subtype"]).strip(),
                "PeakRegion": peak_name,
                "RegionLo": lo,
                "RegionHi": hi,
                "WM1": float(r[col_m1]) if pd.notna(r[col_m1]) else np.nan,
                "WM2": mu2,
                "Sigma": sigma,
            })

    dlong = pd.DataFrame(rows)
    dlong = dlong.replace([np.inf, -np.inf], np.nan).dropna(subset=["WM1", "Sigma"])

    if dlong.empty:
        print("[PlotPeakCentreAndSpread] No usable WM1/WM2 data found.")
        return None

    # optional x normalisation within each peak region
    if normalise_x:
        norm_frames = []
        for peak, g in dlong.groupby("PeakRegion", sort=False):
            g = g.copy()
            vals = g["WM1"].values
            vmin = np.nanmin(vals)
            vmax = np.nanmax(vals)
            if np.isfinite(vmin) and np.isfinite(vmax) and vmax > vmin:
                g["XPlot"] = (vals - vmin) / (vmax - vmin)
                g["SigmaPlot"] = g["Sigma"] / (vmax - vmin)
            else:
                g["XPlot"] = 0.5
                g["SigmaPlot"] = 0.0
            norm_frames.append(g)
        dlong = pd.concat(norm_frames, ignore_index=True)
    else:
        dlong["XPlot"] = dlong["WM1"]
        dlong["SigmaPlot"] = dlong["Sigma"]

    # optional delta from first subtype
    if x_mode == "delta_from_first":
        base = typelist[0]
        shifted = []
        for peak, g in dlong.groupby("PeakRegion", sort=False):
            g = g.copy()
            base_mean = np.nanmean(g.loc[g["Subtype"] == base, "XPlot"])
            g["XPlot"] = g["XPlot"] - base_mean
            shifted.append(g)
        dlong = pd.concat(shifted, ignore_index=True)

    # summarise by subtype within peak region
    summary_rows = []
    for (peak, st), g in dlong.groupby(["PeakRegion", "Subtype"], sort=False):
        xvals = g["XPlot"].astype(float).values
        svals = g["SigmaPlot"].astype(float).values

        summary_rows.append({
            "PeakRegion": peak,
            "Subtype": st,
            "XMean": np.nanmean(xvals),
            "XErr": np.nanstd(xvals, ddof=1) / np.sqrt(np.sum(np.isfinite(xvals))) if np.sum(np.isfinite(xvals)) > 1 else 0.0,
            "SigmaMean": np.nanmean(svals),
            "SigmaErr": np.nanstd(svals, ddof=1) / np.sqrt(np.sum(np.isfinite(svals))) if np.sum(np.isfinite(svals)) > 1 else 0.0,
            "N": np.sum(np.isfinite(xvals))
        })

    dsum = pd.DataFrame(summary_rows)

    peak_order = [name for name, _ in peak_regions if name in dsum["PeakRegion"].unique()]
    ybase = np.arange(len(peak_order))[::-1]

    offset_map = {}
    if len(typelist) == 1:
        offsets = [0.0]
    elif len(typelist) == 2:
        offsets = [-0.12, 0.12]
    elif len(typelist) == 3:
        offsets = [-0.18, 0.0, 0.18]
    else:
        offsets = np.linspace(-0.25, 0.25, len(typelist))
    offset_map = dict(zip(typelist, offsets))

    plt.figure(figsize=figsize)

    # background region bounds
    if show_region_bounds and not normalise_x and x_mode == "absolute":
        for i, (peak_name, (lo, hi)) in enumerate(peak_regions):
            if peak_name not in peak_order:
                continue
            y = ybase[peak_order.index(peak_name)]
            plt.axvspan(lo, hi, ymin=max(0, (y - 0.35) / max(1, len(peak_order)-1 + 0.7)),
                        ymax=min(1, (y + 0.35) / max(1, len(peak_order)-1 + 0.7)),
                        alpha=0.08, color="grey")

    for st in typelist:
        colour = colours.get(st) if colours else None
        linestyle = linestyles.get(st) if linestyles else "-"

        dsub = dsum[dsum["Subtype"] == st].set_index("PeakRegion").reindex(peak_order).reset_index()

        for i, row in dsub.iterrows():
            if pd.isna(row["XMean"]) or pd.isna(row["SigmaMean"]):
                continue

            y = ybase[i] + offset_map[st]
            x = row["XMean"]
            sig = row["SigmaMean"]

            # horizontal spread bar
            plt.plot(
                [x - sig, x + sig],
                [y, y],
                color=colour,
                linestyle=linestyle,
                linewidth=2,
                alpha=0.9
            )

            # centre marker
            plt.plot(
                x, y,
                marker="o",
                color=colour,
                markersize=7,
                label=st if i == 0 else None
            )

    plt.yticks(ybase, peak_order)
    plt.xlabel(
        "Peak centre (WM1)" if (not normalise_x and x_mode == "absolute")
        else ("Normalised peak centre" if normalise_x else f"Peak-centre shift from {typelist[0]}")
    )
    plt.ylabel("Peak region")
    plt.title("Peak centre and spread by region")
    plt.legend()
    plt.tight_layout()
    plt.show()

    return dlong, dsum

def PlotAmideRatios(
    xlsx_path,
    TypestoPlot=("CT", "D7", "D14"),
    amideI_regions=("AmideI_LEFT_1530_1590", "AmideI_MIDDLE_1590_1635", "AmideI_RIGHT_1635_1700"),
    amideIII_region="AmideIII_1410_1500",
    colours=None,
    figsize=(10, 5),
    error_mode="sem",   # "sem" or "std"
    verbose=True
):
    """
    Plot:
      1) Amide I / Amide III peak-height ratio
      2) Amide I / Amide III integrated-area ratio

    Reads sample-level values from the 'Averaged' sheet.
    Assumes metrics were calculated from the sample-averaged spectrum.

    Definitions:
      - Amide I height = max(height of the Amide I sub-regions)
      - Amide I area   = sum(area of the Amide I sub-regions)
      - Ratio = Amide I / Amide III
    """
    xlsx_path = Path(xlsx_path)
    df = pd.read_excel(xlsx_path, sheet_name="Averaged")

    if df.empty:
        print("[PlotAmideRatios] Averaged sheet empty.")
        return None

    typelist = [str(t).strip() for t in TypestoPlot]
    df = df[df["Subtype"].astype(str).isin(typelist)].copy()

    if df.empty:
        print("[PlotAmideRatios] No matching subtypes found.")
        return None

    # required columns
    h_cols_I = [f"{r}_max_intensity_mean" for r in amideI_regions]
    a_cols_I = [f"{r}_area_w_mean" for r in amideI_regions]
    h_col_III = f"{amideIII_region}_max_intensity_mean"
    a_col_III = f"{amideIII_region}_area_w_mean"

    needed = h_cols_I + a_cols_I + [h_col_III, a_col_III]
    missing = [c for c in needed if c not in df.columns]
    if missing:
        print("[PlotAmideRatios] Missing columns:")
        for c in missing:
            print(f"  - {c}")
        return None

    # sample-level ratios
    df["AmideI_height_combined"] = df[h_cols_I].max(axis=1)
    df["AmideI_area_combined"] = df[a_cols_I].sum(axis=1)

    df["HeightRatio"] = df["AmideI_height_combined"] / df[h_col_III]
    df["AreaRatio"] = df["AmideI_area_combined"] / df[a_col_III]

    df = df.replace([np.inf, -np.inf], np.nan)

    if verbose:
        print("\n[PlotAmideRatios] Sample-level ratios")
        print(
            df[[
                "SampleNumber", "Subtype",
                "AmideI_height_combined", h_col_III, "HeightRatio",
                "AmideI_area_combined", a_col_III, "AreaRatio"
            ]].to_string(index=False)
        )

    # summarise by subtype
    rows = []
    for st, g in df.groupby("Subtype", sort=False):
        for metric in ["HeightRatio", "AreaRatio"]:
            vals = g[metric].astype(float).dropna().values
            mean = np.nanmean(vals) if len(vals) else np.nan
            if error_mode == "std":
                err = np.nanstd(vals, ddof=1) if np.sum(np.isfinite(vals)) > 1 else 0.0
            else:
                err = (
                    np.nanstd(vals, ddof=1) / np.sqrt(np.sum(np.isfinite(vals)))
                    if np.sum(np.isfinite(vals)) > 1 else 0.0
                )
            rows.append({
                "Subtype": st,
                "Metric": metric,
                "Mean": mean,
                "Err": err,
                "N": np.sum(np.isfinite(vals))
            })

    dsum = pd.DataFrame(rows)

    fig, axes = plt.subplots(1, 2, figsize=figsize, sharey=False)
    metric_titles = {
        "HeightRatio": "Amide I / Amide III height ratio",
        "AreaRatio": "Amide I / Amide III area ratio"
    }

    for ax, metric in zip(axes, ["HeightRatio", "AreaRatio"]):
        dmet = dsum[dsum["Metric"] == metric].set_index("Subtype").reindex(typelist).reset_index()
        x = np.arange(len(typelist))
        y = dmet["Mean"].values
        yerr = dmet["Err"].values
        bar_colours = [colours.get(st) if colours else None for st in dmet["Subtype"]]

        ax.bar(x, y, color=bar_colours, alpha=0.9)
        ax.errorbar(x, y, yerr=yerr, fmt="none", ecolor="k", elinewidth=1, capsize=3)
        ax.set_xticks(x, dmet["Subtype"])
        ax.set_title(metric_titles[metric])
        ax.set_ylabel("Ratio")
        ax.set_xlabel("Subtype")

    plt.tight_layout()
    plt.show()

    return df, dsum

def PlotAnnotatedAverageSpectra(
    data_dict,
    TypestoPlot,
    region="dermis",
    use_FP=True,
    use_treated=True,
    AveragebyType=True,
    show_error=True,
    error_alpha=0.30,
    err_mode="sem",
    peak_regions=None,
    peak_manifest_df=None,
    component_colours=None,
    xlim=None,
    ylim=None,
    colours=None,
    linestyles=None,
    figsize=(12, 7),
    verbose=False
):
    """
    Plot averaged spectra with:
      - solid black lines at peak-region boundaries
      - coloured dashed lines for biochemical assignments

    Default behaviour is subtype-averaged spectra (AveragebyType=True).
    """

    region = str(region).strip().lower()
    typelist = [str(t).strip() for t in (TypestoPlot if isinstance(TypestoPlot, (list, tuple, set)) else [TypestoPlot])]
    typelist_set = set(typelist)

    reg_short = "FP" if use_FP else "EXT"
    treated_tag = "Treated" if use_treated else "Treated_Full"
    region_key = f"{reg_short}_Spectra_{treated_tag}_{region.capitalize()}"

    # ---------- biochemical assignment lines ----------
    if peak_manifest_df is not None:
        ASSIGNMENTS, assign_colours = BuildAssignmentsFromManifest(
            peak_manifest_df,
            component_colours=component_colours
        )
    else:
        ASSIGNMENTS = {}
        assign_colours = {}
        


    def _get_subtype(sample_data):
        return str(sample_data.get("Subtype", "")).strip()

    def _extract_region_spectra(sample_data):
        sd = sample_data.get(region_key, None)
        return sd if isinstance(sd, dict) and len(sd) else None

    def _align_stack(spectra_dict):
        keys = list(spectra_dict.keys())
        sc0 = spectra_dict[keys[0]]
        x0 = np.asarray(sc0.spectral_axis, float)

        Ys = []
        for k in keys:
            sc = spectra_dict[k]
            x = np.asarray(sc.spectral_axis, float)
            y = np.asarray(sc.spectral_data, float)
            if (x.shape != x0.shape) or (not np.allclose(x, x0, rtol=0, atol=1e-9)):
                y = np.interp(x0, x, y)
            Ys.append(y)

        Y = np.vstack(Ys) if Ys else np.empty((0, x0.size))
        return x0, Y

    def _mean_and_err(Y):
        if Y.size == 0:
            return None, None, 0
        y_mean = np.nanmean(Y, axis=0)
        n_specs = Y.shape[0]
        if n_specs <= 1:
            y_err = np.zeros_like(y_mean)
        else:
            y_std = np.nanstd(Y, axis=0, ddof=1)
            y_err = y_std if err_mode.lower() == "std" else (y_std / np.sqrt(n_specs))
        return y_mean, y_err, n_specs

    curves = []

    if AveragebyType:
        pool = {}
        for sample_num, sample_data in data_dict.items():
            st = _get_subtype(sample_data)
            if st not in typelist_set:
                continue

            sd = _extract_region_spectra(sample_data)
            if sd is None:
                if verbose:
                    print(f"[PlotAnnotatedAverageSpectra] Missing {region_key} for sample {sample_num} ({st})")
                continue

            x0, Y = _align_stack(sd)
            if st not in pool:
                pool[st] = {"x": x0, "Ys": [], "sample_nums": set()}
            else:
                x_ref = pool[st]["x"]
                if (x0.shape != x_ref.shape) or (not np.allclose(x0, x_ref, rtol=0, atol=1e-9)):
                    Y = np.vstack([np.interp(x_ref, x0, row) for row in Y]) if Y.size else np.empty((0, x_ref.size))
                    x0 = x_ref

            pool[st]["Ys"].append(Y)
            pool[st]["sample_nums"].add(str(sample_num))

        for st in typelist:
            if st not in pool:
                continue
            d = pool[st]
            x = d["x"]
            Yall = np.vstack([Y for Y in d["Ys"] if Y.size]) if d["Ys"] else np.empty((0, x.size))
            y_mean, y_err, n_specs = _mean_and_err(Yall)
            if y_mean is None:
                continue
            curves.append({
                "subtype": st,
                "label": f"{st} | n_specs={n_specs} | n_samples={len(d['sample_nums'])}",
                "x": x,
                "y_mean": y_mean,
                "y_err": y_err,
            })
    else:
        for sample_num, sample_data in data_dict.items():
            st = _get_subtype(sample_data)
            if st not in typelist_set:
                continue

            sd = _extract_region_spectra(sample_data)
            if sd is None:
                continue

            x, Y = _align_stack(sd)
            y_mean, y_err, n_specs = _mean_and_err(Y)
            if y_mean is None:
                continue

            curves.append({
                "subtype": st,
                "label": f"{st} | sample {sample_num}",
                "x": x,
                "y_mean": y_mean,
                "y_err": y_err,
            })

    if not curves:
        print(f"[PlotAnnotatedAverageSpectra] No curves to plot (key={region_key}).")
        return None

    plt.figure(figsize=figsize)

    # ---------- spectra ----------
    for c in curves:
        st = c["subtype"]
        x, ym, ye = c["x"], c["y_mean"], c["y_err"]

        colour = colours.get(st) if colours else None
        linestyle = linestyles.get(st) if linestyles else "-"

        if show_error and ye is not None:
            plt.fill_between(x, ym - ye, ym + ye, alpha=error_alpha, linewidth=0, color=colour, zorder=1)

        plt.plot(x, ym, label=st if AveragebyType else c["label"],
                 color=colour, linestyle=linestyle, linewidth=2, zorder=3)

    # ---------- solid peak-region boundaries ----------
    if peak_regions:
        used_label = False
        for _, (x1, x2) in peak_regions:
            plt.axvline(x1, color="k", linestyle="-", alpha=0.35, linewidth=1.2,
                        label="Peak region bounds" if not used_label else None, zorder=0)
            plt.axvline(x2, color="k", linestyle="-", alpha=0.35, linewidth=1.2, zorder=0)
            used_label = True

    # ---------- coloured dashed assignment lines + slanted labels ----------
    ax = plt.gca()
    assignment_labels_used = set()

    # collect all assignments by x-position so overlapping labels can be stacked
    pos_to_components = {}
    for name, xs in ASSIGNMENTS.items():
        for xv in xs:
            pos_to_components.setdefault(float(xv), []).append(name)

    # draw lines and labels
    ymin, ymax = ax.get_ylim()
    yr = ymax - ymin if np.isfinite(ymax - ymin) and ymax > ymin else 1.0

    for xv in sorted(pos_to_components.keys()):
        comps = pos_to_components[xv]

        for j, name in enumerate(comps):
            col = assign_colours.get(name, None)

            # only draw the actual vertical line once per component-position pair
            plt.axvline(
                xv,
                color=col,
                linestyle="--",
                alpha=0.6,
                linewidth=1.0,
                label=name if name not in assignment_labels_used else None,
                zorder=0
            )
            assignment_labels_used.add(name)

            # stacked slanted labels above the line
            y_text = ymax - (0.03 + 0.05 * j) * yr
            ax.text(
                xv,
                y_text,
                name,
                rotation=45,
                rotation_mode="anchor",
                ha="left",
                va="bottom",
                color=col,
                fontsize=8,
                alpha=0.9,
                clip_on=False
            )

    if xlim is not None:
        plt.xlim(xlim)
    if ylim is not None:
        plt.ylim(ylim)

    plt.axhline(0, color="k", lw=1.0, alpha=0.5)
    plt.xlabel("Wavenumber (cm$^{-1}$)")
    plt.ylabel("Intensity (a.u.)")
    plt.title(f"Annotated averaged spectra ({region})")
    plt.legend(fontsize="x-small", ncol=2)
    plt.tight_layout()
    plt.show()

    return curves

